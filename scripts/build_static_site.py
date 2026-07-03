#!/usr/bin/env python
from __future__ import annotations

import datetime
import hashlib
import json
import os
import random
import re
import shutil
import time
from collections import Counter, defaultdict
from html import escape
from pathlib import Path

import markdown


BASE_DIR = Path(__file__).resolve().parent.parent
CONTENT_DIR = BASE_DIR / "content" / "instruments"
OUTPUT_DIR = BASE_DIR / "outputs" / "world-instruments-static"
SITE_BASE_PATH = os.environ.get("SITE_BASE_PATH", "").strip()
SITE_DOMAIN = "https://soundweavers-music.github.io"
_TOTAL_INSTRUMENTS = 0  # set in main()

COUNTRY_COORDS = {
    "非洲": (5.0, 20.0), "非洲／西非": (8.0, -5.0), "非洲／中非": (0.0, 20.0),
    "非洲／東非": (0.0, 35.0), "非洲／南非": (-25.0, 25.0), "非洲／北非": (25.0, 10.0),
    "非洲／辛巴威": (-19.0, 30.0),
    "亞洲": (35.0, 100.0), "亞洲／東亞": (35.0, 115.0), "亞洲／東南亞": (10.0, 105.0),
    "亞洲／南亞": (20.0, 78.0), "亞洲／中亞": (45.0, 65.0), "亞洲／西亞": (30.0, 45.0),
    "歐洲": (50.0, 10.0), "歐洲／西歐": (50.0, 0.0), "歐洲／南歐": (42.0, 12.0),
    "歐洲／北歐": (60.0, 15.0), "歐洲／中歐": (50.0, 10.0), "歐洲／東歐": (52.0, 25.0),
    "美洲": (40.0, -100.0), "美洲／北美": (45.0, -100.0), "美洲／中美": (15.0, -90.0),
    "美洲／南美": (-15.0, -60.0), "美洲／加勒比": (20.0, -75.0),
    "大洋洲": (-25.0, 135.0), "大洋洲／澳洲": (-25.0, 135.0),
    "中東": (28.0, 45.0), "印度": (20.0, 78.0), "中國": (35.0, 105.0),
    "日本": (36.0, 138.0), "韓國": (37.0, 127.5), "臺灣": (23.5, 121.0),
    "台灣": (23.5, 121.0), "印尼": (-5.0, 120.0), "泰國": (15.0, 101.0),
    "越南": (16.0, 108.0), "菲律賓": (12.0, 122.0), "緬甸": (22.0, 96.0),
    "柬埔寨": (12.0, 105.0), "馬來西亞": (4.0, 102.0), "尼泊爾": (28.0, 84.0),
    "蒙古": (46.0, 105.0), "土耳其": (39.0, 35.0), "伊朗": (32.0, 53.0),
    "俄羅斯": (60.0, 40.0), "希臘": (39.0, 22.0), "義大利": (42.0, 12.0),
    "西班牙": (40.0, -3.0), "葡萄牙": (39.5, -8.0), "法國": (46.0, 2.0),
    "德國": (51.0, 10.0), "英國": (55.0, -3.0), "愛爾蘭": (53.0, -8.0),
    "荷蘭": (52.0, 5.0), "比利時": (50.5, 4.5), "瑞士": (47.0, 8.0),
    "奧地利": (47.5, 14.0), "波蘭": (52.0, 20.0), "捷克": (50.0, 15.0),
    "匈牙利": (47.0, 20.0), "羅馬尼亞": (46.0, 25.0), "保加利亞": (43.0, 25.0),
    "挪威": (62.0, 10.0), "瑞典": (62.0, 15.0), "丹麥": (56.0, 10.0),
    "芬蘭": (64.0, 26.0), "冰島": (65.0, -18.0),
    "埃及": (27.0, 30.0), "摩洛哥": (32.0, -6.0), "衣索比亞": (9.0, 38.0),
    "肯亞": (0.0, 38.0), "坦尚尼亞": (-6.0, 35.0), "奈及利亞": (8.0, 8.0),
    "迦納": (8.0, -2.0), "塞內加爾": (14.0, -14.0), "南非": (-30.0, 25.0),
    "澳洲": (-25.0, 135.0), "加拿大": (56.0, -106.0), "美國": (40.0, -100.0),
    "墨西哥": (23.0, -102.0), "巴西": (-14.0, -53.0), "阿根廷": (-38.0, -63.0),
    "祕魯": (-9.0, -75.0), "哥倫比亞": (4.0, -73.0), "古巴": (22.0, -79.0),
}

GLOBAL_KEYWORDS = {"全球", "全球／多地", "跨文化／多地", "全球現代", "多地", "國際"}


def get_region_coords(country_str):
    if not country_str:
        return None
    parts = [p.strip() for p in country_str.replace("／", "/").split("/")]
    for part in parts:
        if part in COUNTRY_COORDS:
            return COUNTRY_COORDS[part]
    for part in parts:
        for key, coord in COUNTRY_COORDS.items():
            if part in key or key in part:
                return coord
    return None


def get_all_region_coords(country_str):
    """Return ALL matching coordinates for multi-region country strings."""
    if not country_str:
        return []
    parts = [p.strip() for p in country_str.replace("／", "/").split("/")]
    found = []
    seen = set()
    for part in parts:
        if part in COUNTRY_COORDS:
            coord = COUNTRY_COORDS[part]
            key = (round(coord[0], 1), round(coord[1], 1))
            if key not in seen:
                seen.add(key)
                found.append(coord)
    for part in parts:
        for key, coord in COUNTRY_COORDS.items():
            if part in key or key in part:
                ckey = (round(coord[0], 1), round(coord[1], 1))
                if ckey not in seen:
                    seen.add(ckey)
                    found.append(coord)
    return found


def normalize_base_path(value):
    if not value or value == "/":
        return ""
    value = value.strip("/")
    return f"/{value}"


SITE_BASE_PATH = normalize_base_path(SITE_BASE_PATH)


def site_url(path):
    path = f"/{path.lstrip('/') }"
    return f"{SITE_BASE_PATH}{path}" or "/"


def resolve_url(page_path, target):
    target = f"/{target.lstrip('/')}"
    if SITE_BASE_PATH:
        return f"{SITE_BASE_PATH}{target}"
    if page_path is None:
        return target
    page_dir = page_path.parent
    asset_path = OUTPUT_DIR / target.lstrip("/")
    url = os.path.relpath(asset_path, page_dir).replace("\\", "/")
    if target.endswith("/") and not url.endswith("/"):
        url += "/"
    if url == ".":
        return "./"
    return url


def safe_external_url(value):
    value = (value or "").strip()
    if value.startswith(("https://", "http://")):
        return value
    return ""


def is_wiki_url(value):
    return bool(value and re.search(r"//(?:[^/]+\.)?(?:wikipedia|wikidata|wikimedia)\.org\b", value, flags=re.IGNORECASE))


def strip_wiki_links(html):
    if not html:
        return html
    link_pattern = re.compile(
        r'<a\b[^>]*href=["\'](?:https?://)?(?:[^/]+\.)?(?:wikipedia|wikidata|wikimedia)\.org[^"\']*["\'][^>]*>(.*?)</a>',
        flags=re.IGNORECASE | re.DOTALL,
    )
    url_pattern = re.compile(
        r"https?://(?:[^/\s]+\.)?(?:wikipedia|wikidata|wikimedia)\.org[^\s<>\"']*",
        flags=re.IGNORECASE,
    )
    while True:
        new_html = link_pattern.sub(lambda m: m.group(1), html)
        if new_html == html:
            break
        html = new_html
    return url_pattern.sub("", html)


def parse_frontmatter(text):
    if not text.startswith("---\n"):
        return {}, text
    _, frontmatter, body = text.split("---", 2)
    data = {}
    for line in frontmatter.splitlines():
        if ":" not in line:
            continue
        key, value = line.split(":", 1)
        value = value.strip()
        if value.startswith('"') and value.endswith('"'):
            value = value[1:-1].replace('\\"', '"').replace("\\\\", "\\")
        data[key.strip()] = value
    return data, body.strip()


def slugify(value):
    # Convert fullwidth parentheses to underscores to differentiate from separators
    value = value.replace("（", "_").replace("）", "")
    slug = re.sub(r"[^a-zA-Z0-9一-鿿_]+", "-", value).strip("-_").lower()
    return slug or "unknown"


def build_map_data(instruments):
    """Return features with lat, lng, name, count, url, samples.
    Multi-region instruments (e.g. '亞洲／歐洲') are counted in ALL matching regions."""
    from collections import defaultdict
    coord_counts = defaultdict(int)
    coord_samples = defaultdict(list)
    coord_country_names = defaultdict(set)

    for item in instruments:
        country = item.get("country", "")
        if not country or country in GLOBAL_KEYWORDS:
            continue
        if any(country.startswith(gk) for gk in GLOBAL_KEYWORDS):
            continue
        all_coords = get_all_region_coords(country)
        if not all_coords:
            continue
        for coords in all_coords:
            key = (round(coords[0], 1), round(coords[1], 1))
            coord_counts[key] += 1
            if len(coord_samples[key]) < 5:
                coord_samples[key].append(item["title"])
            coord_country_names[key].add(country)

    features = []
    for (lat, lng), count in sorted(coord_counts.items(), key=lambda x: -x[1]):
        countries_list = sorted(coord_country_names.get((lat, lng), []))
        primary = max(countries_list, key=len) if countries_list else "未知地區"
        all_urls = [SITE_BASE_PATH + '/countries/' + slugify(c) + '/' for c in countries_list]
        features.append({
            "lat": lat,
            "lng": lng,
            "name": primary,
            "count": count,
            "url": all_urls[0] if all_urls else "{SITE_BASE_PATH}/countries/",
            "urls": all_urls[:10],
            "samples": coord_samples.get((lat, lng), [])[:5],
        })
    return features


def parse_youtube_ids(value):
    if not value:
        return []
    ids = [v.strip() for v in re.split(r"[,\s]+", value.strip()) if v.strip()]
    return [v for v in ids if re.match(r'^[A-Za-z0-9_\-]{11}$', v)]


def remove_empty_headings(html):
    """Remove h2 tags that have no content before the next h2/end."""
    return re.sub(r'(<h2>[^<]+</h2>)\s*(?=<h2>|$)', '', html)


def strip_intro_heading(html):
    """Remove the '介紹' h2 heading tag but keep its content."""
    return re.sub(r'<h2>\s*介紹\s*</h2>', '', html)


def read_instruments():
    instruments = []
    for path in sorted(CONTENT_DIR.glob("*.md")):
        meta, body = parse_frontmatter(path.read_text(encoding="utf-8"))
        # Strip tutorial section from body before conversion (it's extracted separately for the tutorial tab)
        body_without_tutorial = re.sub(
            r"^## 教學\s*\n.*?(?=\n## |\Z)",
            "",
            body.strip(),
            flags=re.DOTALL | re.MULTILINE,
        )
        body_html = markdown.markdown(body_without_tutorial, extensions=["extra", "tables", "fenced_code"], output_format="html5")
        body_html = strip_wiki_links(body_html)
        body_html = remove_empty_headings(body_html)
        body_html = strip_intro_heading(body_html)
        instruments.append(
            {
                "slug": path.stem,
                "title": meta.get("title", path.stem),
                "original_name": meta.get("original_name", ""),
                "category": meta.get("category", "其他"),
                "country": meta.get("country", "待考"),
                "era": meta.get("era", "傳統／年代待考"),
                "sound_class": meta.get("sound_class", ""),
                "hs_class": meta.get("hs_class", ""),
                "family": meta.get("family", ""),
                "playing_method": meta.get("playing_method", ""),
                "body_listening": meta.get("body_listening", ""),
                "soundscape": meta.get("soundscape", ""),
                "region_type": meta.get("region_type", ""),
                "image": meta.get("image", ""),
                "youtube_ids": parse_youtube_ids(meta.get("youtube_ids", "")),
                "instrument_key": meta.get("instrument_key", ""),
                "range": meta.get("range", ""),
                "is_popular": meta.get("is_popular", "").lower() == "true",
                "is_uncommon": meta.get("is_uncommon", "").lower() == "true",
                "site_url": meta.get("site_url", ""),
                "html": body_html,
                "tutorial": "",
            }
        )
        # Extract tutorial section from raw body
        tut_match = re.search(
            r"^## 教學\s*\n(.*?)(?=\n## |\Z)",
            body.strip(),
            re.DOTALL | re.MULTILINE,
        )
        if tut_match:
            instruments[-1]["tutorial"] = markdown.markdown(
                tut_match.group(1).strip(),
                extensions=["extra"],
                output_format="html5",
            )
    return instruments


def write(path, content):
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(content, encoding="utf-8")


def page(title, body, page_path=None, meta_extra="", extra_head="", meta_description="", meta_keywords=""):
    keywords_default = "隔壁織音人,世界樂器,世界樂器百科,世界聲音百科,樂器教學,歌唱教學,錄音後製,吉他教學,鋼琴教學,基礎樂理,音樂知識,民族樂器,傳統樂器,電子樂器,打擊樂器,管樂器,弦樂器,鍵盤樂器"
    kw = meta_keywords or keywords_default
    desc = meta_description or "世界聲音百科 by 隔壁織音人 — 收錄世界各國樂器、人聲歌唱教學、錄音後製知識與基礎樂理。從傳統民族樂器到現代電子樂器，提供樂器介紹、聆賞示範、演奏教學與文化背景。循著聲音，走進不同文化的現場。"
    csp = (
        "default-src 'self'; "
        "img-src 'self' https: data:; "
        "style-src 'self' 'unsafe-inline' https://unpkg.com https://cdnjs.cloudflare.com; "
        "script-src 'self' 'unsafe-inline' https://unpkg.com https://cdnjs.cloudflare.com https://busuanzi.ibruce.info https://pagead2.googlesyndication.com; "
        "connect-src 'self' https://busuanzi.ibruce.info; "
        "frame-src https://www.youtube-nocookie.com https://www.youtube.com; "
        "base-uri 'self'; form-action 'none'; object-src 'none'"
    )
    raw_canon = resolve_url(page_path, "/") if page_path else "/"
    clean_canon = raw_canon.replace("./", "/").replace("../", "/").rstrip("/") or "/"
    canonical_url = f"https://soundweavers-music.github.io{clean_canon}".rstrip(".")
    og_image = "https://yt3.googleusercontent.com/6nBZ7RVoXGMH2fuMPWiju_tpAET9D-qVkOhg1HjGqh8m9EaO-u9wO_oHVA12Sy0DzoKn7mGVmA=w1707-fcrop64=1,00005a57ffffa5a8-k-c0xffffffff-no-nd-rj"
    jsonld = f'''{{
  "@context": "https://schema.org",
  "@type": "WebSite",
  "name": "世界聲音百科",
  "alternateName": "World Sound Encyclopedia",
  "url": "https://soundweavers-music.github.io/",
  "description": "{escape(desc)}",
  "author": {{
    "@type": "Person",
    "name": "隔壁織音人",
    "url": "https://www.youtube.com/@NextDoorSoundWeavers/"
  }},
  "potentialAction": {{
    "@type": "SearchAction",
    "target": {{
      "@type": "EntryPoint",
      "urlTemplate": "https://soundweavers-music.github.io/?q={{search_term_string}}"
    }},
    "query-input": "required name=search_term_string"
  }},
  "sameAs": [
    "https://www.youtube.com/@NextDoorSoundWeavers/"
  ]
}}'''
    seo_tags = f'''<meta name="description" content="{escape(desc)}">
<meta name="keywords" content="{escape(kw)}">
<link rel="canonical" href="{canonical_url}">
<script type="application/ld+json">{jsonld}</script>'''''
    if "og:title" not in meta_extra:
        seo_tags += f'''
<meta property="og:title" content="{escape(title) if "|" in title or "｜" in title else escape(title) + "｜世界聲音百科"}">
<meta property="og:description" content="{escape(desc)}">
<meta property="og:image" content="{og_image}">
<meta property="og:type" content="website">'''
    _dm_head = '<script>!function(){try{var t=localStorage.getItem("theme");if(t)document.documentElement.setAttribute("data-theme",t);else document.documentElement.setAttribute("data-theme","nextdoor")}catch(e){}}()</script>'
    _dm_foot = '<script>(function(){var t=document.getElementById("theme-toggle"),d=document.documentElement;var I={nextdoor:"🎋",light:"🌤",dark:"🌙"};function s(m){d.setAttribute("data-theme",m);if(t)t.textContent=I[m]||"🎋";try{localStorage.setItem("theme",m)}catch(e){}}var v=(localStorage.getItem("theme")||"nextdoor");s(v);if(t)t.addEventListener("click",function(){var c=d.getAttribute("data-theme");s(c==="nextdoor"?"light":c==="light"?"dark":"nextdoor")});})();</script>'

    return f"""<!doctype html>
<html lang="zh-Hant">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <meta name="referrer" content="no-referrer-when-downgrade">
  <meta http-equiv="Content-Security-Policy" content="{csp}">
  <meta name="google-site-verification" content="AzedQ-PxUmSW7_0jyEHmHCKgN2nIK0Bio5d6LCsJTtE">
  <script async src="https://pagead2.googlesyndication.com/pagead/js/adsbygoogle.js?client=ca-pub-6561686484716387" crossorigin="anonymous"></script>
  <title>{escape(title) if "|" in title or "｜" in title else escape(title) + "｜世界聲音百科"}</title>
  {seo_tags}
  {meta_extra}
  {_dm_head}
  <link rel="stylesheet" href="{resolve_url(page_path, '/assets/site.css')}">
  {extra_head}
</head>
<body>
  <header class="site-header">
    <a class="brand" href="{resolve_url(page_path, '/')}">🌍 世界聲音百科</a>
    <nav>
      <div class="nav-dropdown">
        <a href="{resolve_url(page_path, '/instruments/')}" class="dropdown-trigger">全部樂器</a>
        <div class="dropdown-menu">
          <a href="{resolve_url(page_path, '/categories/')}">分類</a>
          <a href="#" id="random-nav-link" class="random-link">隨選</a>
          <a href="{resolve_url(page_path, '/popular/')}">熱門</a>
          <a href="{resolve_url(page_path, '/uncommon/')}">冷門</a>
          <a href="{resolve_url(page_path, '/countries/')}">國家</a>
          <a href="{resolve_url(page_path, '/eras/')}">年代</a>
          <a href="{resolve_url(page_path, '/map/')}">地圖</a>
        </div>
      </div>
      <div class="nav-dropdown">
        <a href="{resolve_url(page_path, '/sound-journey/')}" class="dropdown-trigger">聲音旅圖</a>
        <div class="dropdown-menu">
          <a href="{resolve_url(page_path, '/sound-journey/')}">旅圖門口</a>
          <a href="{resolve_url(page_path, '/sound-journey/1/')}">旅圖一｜氣息離開身體</a>
          <a href="{resolve_url(page_path, '/sound-journey/2/')}">旅圖二｜懷裡與大地的弦</a>
          <a href="{resolve_url(page_path, '/sound-journey/3/')}">旅圖三｜天空裡拉長的弦</a>
          <a href="{resolve_url(page_path, '/sound-journey/4/')}">旅圖四｜地心與手邊的微光</a>
          <a href="{resolve_url(page_path, '/sound-journey/5/')}">旅圖五｜城市與星塵的節奏</a>
          <a href="{resolve_url(page_path, '/sound-journey/6/')}">旅圖六｜按鍵打開房間，眾聲走向廣場</a>
          <a href="{resolve_url(page_path, '/sound-journey/all/')}">全部文章</a>
        </div>
      </div>
      <div class="nav-dropdown">
        <a href="{resolve_url(page_path, '/vocal/')}" class="dropdown-trigger">人聲與歌唱</a>
        <div class="dropdown-menu">
          <a href="{resolve_url(page_path, '/vocal/')}">課程總覽</a>
          <a href="{resolve_url(page_path, '/vocal/1/')}">初階篇</a>
          <a href="{resolve_url(page_path, '/vocal/16/')}">進階篇</a>
        </div>
      </div>
      <div class="nav-dropdown">
        <a href="{resolve_url(page_path, '/digitalmusic/')}" class="dropdown-trigger">錄音後製</a>
        <div class="dropdown-menu">
          <a href="{resolve_url(page_path, '/digitalmusic/')}">課程總覽</a>
          <a href="{resolve_url(page_path, '/digitalmusic/1/')}">基礎篇</a>
          <a href="{resolve_url(page_path, '/digitalmusic/16/')}">進階篇</a>
        </div>
      </div>
      <div class="nav-dropdown">
        <a href="{resolve_url(page_path, '/theory/')}" class="dropdown-trigger">樂理基礎</a>
        <div class="dropdown-menu">
          <a href="{resolve_url(page_path, '/theory/')}">課程總覽</a>
          <a href="{resolve_url(page_path, '/theory/0/')}">前言</a>
          <a href="{resolve_url(page_path, '/theory/1/')}">階段一</a>
          <a href="{resolve_url(page_path, '/theory/2/')}">階段二</a>
          <a href="{resolve_url(page_path, '/theory/3/')}">階段三</a>
          <a href="{resolve_url(page_path, '/theory/4/')}">階段四</a>
          <a href="{resolve_url(page_path, '/theory/5/')}">階段五</a>
        </div>
      </div>
      <a href="{resolve_url(page_path, '/experience/')}">🎹 體驗</a>
      <a href="{resolve_url(page_path, '/about/')}">關於</a>
      <a href="{resolve_url(page_path, '/contact/')}">聯絡我們</a>
    </nav>
    <button id="theme-toggle" class="theme-toggle" aria-label="切換色調">🎋</button>
  </header>
  {body}
  <footer class="site-footer">
    <div class="footer-inner">
      <span>世界聲音百科 — 世界樂器、人聲與音樂文化的旅圖</span>
      <span>作者：<a href="https://www.youtube.com/@NextDoorSoundWeavers/" target="_blank" rel="noopener">隔壁織音人</a></span>
      <nav class="footer-nav">
        <a href="{resolve_url(page_path, '/')}">首頁</a>
        <a href="{resolve_url(page_path, '/sound-journey/')}">聲音旅圖</a>
        <a href="{resolve_url(page_path, '/vocal/')}">人聲與歌唱</a>
	      <a href="{resolve_url(page_path, '/digitalmusic/')}">錄音後製</a>
        <a href="{resolve_url(page_path, '/categories/')}">分類</a>
        <a href="{resolve_url(page_path, '/countries/')}">國家</a>
        <a href="{resolve_url(page_path, '/popular/')}">熱門</a>
        <a href="{resolve_url(page_path, '/uncommon/')}">冷門</a>
        <a href="{resolve_url(page_path, '/theory/')}">樂理基礎</a>
        <a href="{resolve_url(page_path, '/contact/')}">聯絡我們</a>
        <a href="https://www.youtube.com/@NextDoorSoundWeavers/" target="_blank" rel="noopener">訂閱 YouTube</a>
      <span class="visit-counter">總瀏覽次數：<span id="busuanzi_value_site_pv"></span> ｜今日訪客：<span id="busuanzi_value_site_uv"></span></span>
      </nav>
    </div>
  </footer>
  <button id="back-top" class="back-top" aria-label="回頂部">↑</button>
  <script src="{resolve_url(page_path, '/assets/search.js')}"></script>
    <script src="{resolve_url(page_path, '/assets/random-instrument.js')}"></script>
  <script async src="https://busuanzi.ibruce.info/busuanzi/2.3/busuanzi.pure.mini.js"></script>
  {_dm_foot}
</body>
</html>
"""


def card(instrument, page_path=None):
    img = safe_external_url(instrument.get("image", ""))
    img_html = f'<img class="card-thumb" src="{img}" alt="" loading="lazy" onerror="this.style.display=\'none\'">' if img else '<div class="card-thumb card-thumb--empty" aria-hidden="true">♩</div>'
    return f"""<a class="instrument-card" href="{resolve_url(page_path, '/instruments/' + instrument['slug'] + '/')}">
      {img_html}
      <div class="card-body">
        <span class="card-cat">{escape(instrument['category'])}</span>
        <strong class="card-title">{escape(instrument['title'])}</strong>
        {f'<span class="card-orig">{escape(instrument["original_name"])}</span>' if instrument.get("original_name") and instrument["original_name"] != instrument["title"] else ""}
        <span class="card-meta">{escape(instrument['country'])}</span>
      </div>
    </a>"""


def list_page(title, instruments, page_path=None, meta_description=""):
    desc = meta_description or f"{title} — {len(instruments)} 件樂器的完整列表，包含樂器介紹、聆賞示範、演奏教學與文化背景。世界聲音百科 by 隔壁織音人。"
    cards = "\n".join(card(item, page_path) for item in instruments) or '<p class="empty">目前沒有資料。</p>'
    return page(
        title,
        f"""
        <main class="page">
          <section class="compact-hero">
            <p class="eyebrow">Browse</p>
            <h1>{escape(title)}</h1>
            <p class="lead">{len(instruments)} 件樂器</p>
          </section>
          <div class="instrument-grid">{cards}</div>
        </main>
        """,
        page_path,
        meta_description=desc,
    )


def build_instruments_list_page(instruments):
    """Build the /instruments/ page with view-switch: filter dropdown + category cards."""
    page_path = OUTPUT_DIR / "instruments" / "index.html"
    by_cat = defaultdict(list)
    for item in instruments:
        by_cat[item["category"]].append(item)
    category_links = "".join(
        f'<a class="facet-card" href="../../categories/{slugify(name)}/"><strong>{escape(name)}</strong><span>{len(items)} 筆</span></a>'
        for name, items in sorted(by_cat.items())
    )
    cards = "\n".join(card(item, page_path) for item in instruments) or '<p class="empty">目前沒有資料。</p>'
    body = f"""
    <main class="page">
      <section class="compact-hero">
        <p class="eyebrow">All Instruments</p>
        <h1>全部樂器</h1>
        <p class="lead">{len(instruments)} 件樂器</p>
        <div class="search-panel" style="max-width:520px;margin-top:16px;">
          <input id="site-search" type="search" placeholder="搜尋中文名、英文名、分類、國家或年代…" autocomplete="off" spellcheck="false">
        </div>
        <div id="search-results" class="search-results"></div>
      </section>

      <section class="view-switch" aria-label="瀏覽模式">
        <button id="mode-dropdown" class="is-active" type="button">🔍 篩選瀏覽</button>
        <button id="mode-cards" type="button">📂 分類卡片</button>
      </section>

      <div id="dropdown-mode" class="browse-mode">
        <section class="section">
          <div class="section-heading"><h2>篩選樂器</h2><span id="dropdown-count" class="section-note"></span></div>
          <div class="dropdown-browser">
            <label>
              <span>分類</span>
              <select id="filter-category"><option value="">全部分類</option></select>
            </label>
            <label>
              <span>國家/地區</span>
              <select id="filter-country"><option value="">全部國家/地區</option></select>
            </label>
            <label>
              <span>年代</span>
              <select id="filter-era"><option value="">全部年代</option></select>
            </label>
            <label>
              <span>發聲方式</span>
              <select id="filter-sound-class"><option value="">全部發聲</option></select>
            </label>
            <button id="filter-reset" type="button">✕ 重設</button>
          </div>
          <div id="dropdown-results" class="dropdown-results"></div>
        </section>
      </div>

      <div id="card-mode" class="browse-mode" hidden>
        <section class="section">
          <div class="section-heading"><h2>分類瀏覽</h2></div>
          <div class="facet-grid">{category_links}</div>
        </section>
        <section class="section">
          <div class="section-heading"><h2>全部樂器</h2><span class="section-note">{len(instruments)} 件</span></div>
          <div class="instrument-grid">{cards}</div>
        </section>
      </div>
    </main>
    """
    write(page_path, page("全部樂器", body, page_path, meta_description="世界樂器百科 — 收錄世界各國 897 件傳統與現代樂器，可按分類、國家、年代、發聲方式篩選。提供樂器介紹、音色描述、歷史背景、演奏教學與 YouTube 聆賞示範。由隔壁織音人整理。"))


def build_index(instruments):
    index_path = OUTPUT_DIR / "index.html"
    categories = Counter(item["category"] for item in instruments)
    countries = Counter(item["country"] for item in instruments)
    eras = Counter(item["era"] for item in instruments)
    category_links = "".join(
        f'<a class="facet-card" href="{resolve_url(index_path, f"/categories/{slugify(name)}/")}"><strong>{escape(name)}</strong><span>{count} 筆</span></a>'
        for name, count in categories.most_common()
    )
    # Pick 12 featured instruments: 2 per category, image preferred, deterministic
    by_cat = defaultdict(list)
    for item in instruments:
        by_cat[item["category"]].append(item)
    featured = []
    rng = random.Random(42)
    for cat_items in by_cat.values():
        with_img = [i for i in cat_items if i.get("image")]
        pool = with_img if with_img else cat_items
        shuffled = list(pool)
        rng.shuffle(shuffled)
        featured.extend(shuffled[:2])
    rng.shuffle(featured)
    sample_cards = "\n".join(card(item, index_path) for item in featured[:12])

    # Popular/uncommon counts
    pop_count = sum(1 for i in instruments if i.get("is_popular"))
    un_count = sum(1 for i in instruments if i.get("is_uncommon"))

    body = f"""
    <main class="page">
      <section class="hero">
        <p class="eyebrow">World Musical Instruments Encyclopedia</p>
        <h1>世界聲音百科</h1>
        <p class="lead hero-lead">收錄來自世界各地的傳統與現代樂器，探索人類音樂的多元面貌。</p>
        <div class="search-panel">
          <input id="site-search" type="search" placeholder="搜尋中文名、英文名、分類、國家或年代…" autocomplete="off" spellcheck="false">
        </div>
        <div id="search-results" class="search-results"></div>
      </section>

      <section class="stats">
        <div class="stat-item"><strong>{len(instruments)}</strong><span>樂器條目</span></div>
        <div class="stat-item"><strong>{len(categories)}</strong><span>分類</span></div>
        <div class="stat-item"><strong>{len(countries)}</strong><span>國家/地區</span></div>
        <div class="stat-item"><strong>{len(eras)}</strong><span>年代</span></div>
      </section>


      <section class="section">
        <div class="section-heading"><h2>精選分類</h2></div>
        <div class="featured-links">
          <a class="featured-card hot" href="{resolve_url(index_path, '/popular/')}">
            <strong>熱門樂器</strong>
            <span>{pop_count} 件熱門樂器，探索世界知名樂器</span>
          </a>
          <a class="featured-card cold" href="{resolve_url(index_path, '/uncommon/')}">
            <strong>冷門樂器</strong>
            <span>{un_count} 件冷門樂器，發掘稀有珍品</span>
          </a>
          <a class="featured-card random" href="#" id="random-link-home">
            <strong>隨選樂器</strong>
            <span>每次點選都隨機產生一個樂器，驚喜不斷</span>
          </a>
        </div>
      </section>

      <section class="view-switch" aria-label="瀏覽模式">
        <button id="mode-dropdown" class="is-active" type="button">🔍 篩選瀏覽</button>
        <button id="mode-cards" type="button">📂 分類卡片</button>
      </section>

      <div id="dropdown-mode" class="browse-mode">
        <section class="section">
          <div class="section-heading"><h2>篩選樂器</h2><span id="dropdown-count" class="section-note"></span></div>
          <div class="dropdown-browser">
            <label>
              <span>分類</span>
              <select id="filter-category"><option value="">全部分類</option></select>
            </label>
            <label>
              <span>國家/地區</span>
              <select id="filter-country"><option value="">全部國家/地區</option></select>
            </label>
            <label>
              <span>年代</span>
              <select id="filter-era"><option value="">全部年代</option></select>
            </label>
            <label>
              <span>發聲方式</span>
              <select id="filter-sound-class"><option value="">全部發聲</option></select>
            </label>
            <button id="filter-reset" type="button">✕ 重設</button>
          </div>
          <div id="dropdown-results" class="dropdown-results"></div>
        </section>
      </div>

      <div id="card-mode" class="browse-mode" hidden>
        <section class="section">
          <div class="section-heading"><h2>分類瀏覽</h2><a href="{resolve_url(index_path, '/categories/')}">全部分類 →</a></div>
          <div class="facet-grid">{category_links}</div>
        </section>

        <section class="section">
          <div class="section-heading"><h2>精選樂器</h2><a href="{resolve_url(index_path, '/instruments/')}">查看全部 →</a></div>
          <div class="instrument-grid">{sample_cards}</div>
        </section>
      </div>
    </main>
    """
    write(index_path, page("世界聲音百科 | 隔壁織音人", body, index_path, meta_description="世界聲音百科 by 隔壁織音人 — 收錄世界各國傳統與現代樂器（卡林巴、手碟、鋼琴、吉他、小提琴等）、人聲歌唱教學、錄音後製知識與基礎樂理。探索樂器分類、國家地區、年代與音色，適合音樂創作者、教育工作者與樂器愛好者。"))

    # Write map data as JSON for the map page
    map_features = build_map_data(instruments)
    map_json_path = OUTPUT_DIR / "assets" / "map-data.json"
    map_json_path.write_text(json.dumps(map_features, ensure_ascii=False, indent=2), encoding="utf-8")

    # Write random-instrument JS for the nav link and homepage card
    slugs_json = json.dumps([item["slug"] for item in instruments], ensure_ascii=False)
    rand_js = f"""
function goRandom() {{
  var slugs = {slugs_json};
  var idx = Math.floor(Math.random() * slugs.length);
  window.location.href = '{SITE_BASE_PATH}/instruments/' + slugs[idx] + '/';
}}
document.getElementById('random-link-home')?.addEventListener('click', function(e) {{ e.preventDefault(); goRandom(); }});
document.getElementById('random-nav-link')?.addEventListener('click', function(e) {{ e.preventDefault(); goRandom(); }});
"""
    write(OUTPUT_DIR / "assets" / "random-instrument.js", rand_js.strip() + "\n")


def meta_row(label, value):
    if not value:
        return ""
    return f'<div class="meta-item"><dt>{escape(label)}</dt><dd>{escape(value)}</dd></div>'


def build_youtube_grid(youtube_ids):
    """Return the yt-grid div HTML for injecting into the 聆聽示範 section."""
    if not youtube_ids:
        return ""
    iframes = []
    for vid_id in youtube_ids[:2]:
        iframes.append(
            f'<div class="yt-embed"><iframe src="https://www.youtube-nocookie.com/embed/{escape(vid_id)}" '
            f'title="YouTube video player" frameborder="0" '
            f'allow="accelerometer; autoplay; clipboard-write; encrypted-media; gyroscope; picture-in-picture" '
            f'allowfullscreen loading="lazy"></iframe></div>'
        )
    return f'<div class="yt-grid">{"".join(iframes)}</div>'


def inject_youtube_into_body(body_html, youtube_ids):
    """Inject YouTube iframes after the <h2>聆聽示範</h2> heading in body HTML."""
    grid_html = build_youtube_grid(youtube_ids)
    if not grid_html:
        return body_html
    return body_html.replace(
        '<h2>聆聽示範</h2>',
        f'<h2>聆聽示範</h2>{grid_html}',
        1,
    )


def build_detail_pages(instruments):
    by_category = defaultdict(list)
    for item in instruments:
        by_category[item["category"]].append(item)

    for item in instruments:
        meta_fields = [
            ("分類", item["category"]),
            ("國家／地區", item["country"]),
            ("年代", item["era"]),
            ("發聲原理", item["sound_class"]),
            ("樂器調性", item.get("instrument_key", "")),
            ("音域範圍", item.get("range", "")),
            ("HS 分類", item["hs_class"]),
            ("樂器家族", item["family"]),
            ("演奏方式", item["playing_method"]),
            ("身體聆聽", item["body_listening"]),
            ("地區類型", item["region_type"]),
        ]
        meta_grid = "".join(meta_row(label, val) for label, val in meta_fields if val)
        orig = f'<p class="original-name">{escape(item["original_name"])}</p>' if item["original_name"] and item["original_name"] != item["title"] else ""
        soundscape_val = item.get("soundscape", "")
        soundscape_html = f'<p class="soundscape-tag">{escape(soundscape_val)}</p>' if soundscape_val else ""
        img_url = safe_external_url(item.get("image", ""))
        img_html = f'<img class="instrument-image" src="{img_url}" alt="{escape(item["title"])}" loading="lazy" onerror="this.style.display=\'none\'">' if img_url else ""
        img_credit_html = '<p class="image-credit">圖片來源：Wikimedia Commons</p>' if img_url else ""
        header_class = "instrument-header has-image" if img_url else "instrument-header"
        # Popularity badges
        badges = ""
        if item.get("is_popular"):
            badges += '<span class="badge badge-hot">熱門</span>'
        if item.get("is_uncommon"):
            badges += '<span class="badge badge-cold">冷門</span>'
        # Inject YouTube iframes into the 聆聽示範 section in the article body
        body_html = inject_youtube_into_body(item['html'], item.get("youtube_ids", []))
        # Extract plain-text description from first paragraph of HTML body
        desc_match = re.search(r'<p>(.*?)</p>', body_html, re.DOTALL)
        desc_text = re.sub(r'<[^>]+>', '', desc_match.group(1))[:160] if desc_match else ""
        keywords = f"{item['title']},{item['original_name']},{item['category']},{item['country']},樂器教學,世界樂器,樂器介紹"
        article_tags = "\n".join(
            f'<meta property="article:tag" content="{escape(tag)}">'
            for tag in [item['category'], item['country'], item.get('era', ''), item.get('sound_class', '')]
            if tag
        )
        og_tags = "\n".join(filter(None, [
            f'<meta property="og:title" content="{escape(item["title"])}｜世界聲音百科">',
            f'<meta property="og:description" content="{escape(desc_text)}">' if desc_text else "",
            f'<meta property="og:image" content="{img_url}">' if img_url else "",
            f'<meta property="og:type" content="article">',
            article_tags if article_tags else "",
        ]))
        related_pool = [i for i in by_category[item["category"]] if i["slug"] != item["slug"]]
        seed = int(hashlib.md5(item["slug"].encode()).hexdigest(), 16)
        random.Random(seed).shuffle(related_pool)
        related = related_pool[:4]
        related_html = ""
        if related:
            related_cards = "\n".join(card(r) for r in related)
            related_html = f'<section class="related-section"><h2 class="related-heading">同類樂器</h2><div class="instrument-grid">{related_cards}</div></section>'
        body = f"""
        <main class="instrument-page">
          <nav class="breadcrumb">
            <a href="../../">← 返回篩選</a> <span class="sep">/</span>
            <span>{escape(item['title'])}</span>
          </nav>
          <header class="{header_class}">
            <div class="header-text">
              <p class="eyebrow">{escape(item['category'])}</p>
              <h1>{escape(item['title'])}</h1>
              {orig}
              {soundscape_html}
              <p class="badge-row">{badges}</p>
            </div>
            {f'<div class="header-image">{img_html}{img_credit_html}</div>' if img_url else ""}
          </header>
          {"<dl class='meta-grid'>" + meta_grid + "</dl>" if meta_grid else ""}
          {f'''
          <div class="tab-bar">
            <button class="tab-btn is-active" data-tab="intro">介紹</button>
            <button class="tab-btn" data-tab="tutorial">教學</button>
          </div>
          <div id="tab-intro" class="tab-pane is-active"><article class="markdown-body">{body_html}</article></div>
          <div id="tab-tutorial" class="tab-pane"><article class="markdown-body">{item["tutorial"]}</article></div>
          '''}
          {related_html}
        </main>
        """
        write(OUTPUT_DIR / "instruments" / item["slug"] / "index.html", page(item["title"], body, meta_extra=og_tags, meta_description=desc_text, meta_keywords=keywords))


def build_facet_pages(instruments, field, folder, title):
    grouped = defaultdict(list)
    for item in instruments:
        if item.get(field):
            grouped[item[field]].append(item)

    facet_cards = "".join(
        f'<a class="facet-card" href="{site_url(f"/{folder}/{slugify(name)}/")}"><strong>{escape(name)}</strong><span>{len(items)} 筆</span></a>'
        for name, items in sorted(grouped.items())
    )
    desc_map = {"category": "世界樂器分類一覽", "country": "世界樂器國家地區一覽", "era": "世界樂器年代一覽", "sound_class": "世界樂器發聲方式一覽"}
    browse_desc = desc_map.get(field, f"{title}一覽")
    write(
        OUTPUT_DIR / folder / "index.html",
        page(title, f'<main class="page"><section class="compact-hero"><h1>{escape(title)}</h1></section><div class="facet-grid">{facet_cards}</div></main>', meta_description=f"{browse_desc} — 探索來自各{field}的樂器，包含樂器介紹、聆賞示範與文化背景。世界聲音百科 by 隔壁織音人。"),
    )
    for name, items in grouped.items():
        write(OUTPUT_DIR / folder / slugify(name) / "index.html", list_page(name, sorted(items, key=lambda item: item["title"])))


def build_assets(instruments):
    css = """
/* ── Reset & tokens ─────────────────────────────────────────── */
:root {
  --ink: #1a2332;
  --ink2: #344054;
  --muted: #667085;
  --line: #e4e7ec;
  --surface: #fff;
  --soft: #f8fafc;
  --accent: #0d766b;
  --accent2: #0a5c53;
  --blue: #1d4ed8;
  --radius: 10px;
  --shadow: 0 1px 3px rgba(0,0,0,.08), 0 1px 2px rgba(0,0,0,.06);
  --shadow-md: 0 4px 6px -1px rgba(0,0,0,.07), 0 2px 4px -2px rgba(0,0,0,.07);
}
[data-theme="dark"] {
  --ink: #e2e8f0;
  --ink2: #cbd5e1;
  --muted: #94a3b8;
  --line: #334155;
  --surface: #1e293b;
  --soft: #0f172a;
  --accent: #5eead4;
  --accent2: #14b8a6;
  --blue: #60a5fa;
  --shadow: 0 1px 3px rgba(0,0,0,.3), 0 1px 2px rgba(0,0,0,.25);
  --shadow-md: 0 4px 6px -1px rgba(0,0,0,.35), 0 2px 4px -2px rgba(0,0,0,.3);
}
[data-theme="nextdoor"] {
  --ink: #3D3229;
  --ink2: #5D4E41;
  --muted: #9C8F87;
  --line: #E5DDD6;
  --surface: #FFFFFF;
  --soft: #F7F3EE;
  --accent: #C4956A;
  --accent2: #A67C52;
  --blue: #6B8FBF;
  --shadow: 0 1px 3px rgba(0,0,0,.05), 0 1px 2px rgba(0,0,0,.04);
  --shadow-md: 0 4px 6px -1px rgba(0,0,0,.04), 0 2px 4px -2px rgba(0,0,0,.04);
}
*, *::before, *::after { box-sizing: border-box; }
body { margin:0; color:var(--ink); background:var(--soft); font-family:-apple-system,BlinkMacSystemFont,"Segoe UI","Noto Sans TC",sans-serif; line-height:1.6; }
a { color:inherit; }
img { max-width:100%; }

/* ── Site header ─────────────────────────────────────────────── */
.site-header {
  display:flex; justify-content:space-between; gap:20px; align-items:center;
  padding:14px 28px; border-bottom:1px solid var(--line);
  background:var(--surface); backdrop-filter:blur(8px);
  position:sticky; top:0; z-index:100;
  box-shadow: 0 1px 0 var(--line);
}
.brand { font-weight:800; font-size:17px; text-decoration:none; color:var(--accent); letter-spacing:-.3px; }
.site-header nav { display:flex; gap:4px; }
.site-header nav a { text-decoration:none; color:var(--muted); font-size:14px; font-weight:500; padding:6px 10px; border-radius:6px; transition:color .15s,background .15s; }
.site-header nav a:hover { color:var(--ink); background:var(--soft); }

/* ── Theme toggle ─────────────────────────────────────────────── */
.theme-toggle { background:none; border:1px solid var(--line); border-radius:6px; cursor:pointer; font-size:16px; line-height:1; padding:5px 8px; color:var(--ink2); transition:border-color .15s; flex-shrink:0; margin-left:4px; }
.theme-toggle:hover { border-color:var(--accent); }

/* ── Dropdown nav ─────────────────────────────────────────────── */
.nav-dropdown { position:relative; display:flex; align-items:stretch; }
.nav-dropdown .dropdown-trigger { cursor:pointer; display:flex; align-items:center; }
.dropdown-menu {
  display:none; position:absolute; top:100%; left:0; z-index:200;
  min-width:140px; background:var(--surface); border:1px solid var(--line);
  border-radius:8px; box-shadow:var(--shadow-md); padding:4px 0;
}
.nav-dropdown:hover .dropdown-menu,
.nav-dropdown .dropdown-menu:hover { display:block; }
.dropdown-menu a {
  display:block; padding:8px 16px; font-size:13px; color:var(--ink2);
  text-decoration:none; border-radius:0; white-space:nowrap;
}
.dropdown-menu a:hover { background:var(--soft); color:var(--accent); }

/* ── Page layout ─────────────────────────────────────────────── */
.page,.instrument-page { max-width:1160px; margin:0 auto; padding:36px 24px 80px; }

/* ── Hero ────────────────────────────────────────────────────── */
.hero { padding:60px 0 40px; }
.compact-hero { padding:32px 0 28px; }
.eyebrow { color:var(--accent); font-size:12px; font-weight:700; margin:0 0 10px; text-transform:uppercase; letter-spacing:.08em; }
h1 { font-size:clamp(32px,5vw,48px); line-height:1.1; margin:0 0 16px; font-weight:800; letter-spacing:-.5px; }
h2 { margin:0; font-weight:700; }
.lead { color:var(--muted); line-height:1.7; margin:0 0 8px; }
.hero-lead { max-width:520px; font-size:17px; margin:0 0 32px; }
.empty { color:var(--muted); }

/* ── Search ──────────────────────────────────────────────────── */
.search-panel { max-width:580px; }
.search-panel input {
  width:100%; height:52px; border:2px solid var(--line); border-radius:var(--radius);
  padding:0 18px; font-size:16px; background:var(--surface); color:var(--ink);
  transition:border-color .2s, box-shadow .2s;
}
.search-panel input:focus { outline:none; border-color:var(--accent); box-shadow:0 0 0 3px rgba(13,118,107,.12); }
.search-results { margin-top:10px; display:grid; gap:6px; max-width:580px; }

/* ── Stats ───────────────────────────────────────────────────── */
.stats { display:grid; grid-template-columns:repeat(4,minmax(120px,1fr)); gap:12px; margin:0 0 40px; }
.stat-item { border:1px solid var(--line); background:var(--surface); border-radius:var(--radius); padding:20px; text-align:center; box-shadow:var(--shadow); }
.stat-item strong { display:block; font-size:32px; font-weight:800; color:var(--accent); line-height:1.1; }
.stat-item span { color:var(--muted); font-size:13px; }

/* ── View switch ─────────────────────────────────────────────── */
.view-switch { display:flex; flex-wrap:wrap; gap:8px; margin:0 0 4px; }
.view-switch button {
  height:40px; border:2px solid var(--line); border-radius:8px;
  padding:0 18px; background:var(--surface); color:var(--muted);
  font-weight:700; font-size:14px; cursor:pointer; transition:all .15s;
}
.view-switch button:hover { border-color:var(--accent); color:var(--accent); }
.view-switch button.is-active { border-color:var(--accent); background:var(--accent); color:#fff; }
.browse-mode[hidden] { display:none !important; }

/* ── Section ─────────────────────────────────────────────────── */
.section { margin-top:40px; }
.section-heading { display:flex; justify-content:space-between; align-items:center; margin-bottom:18px; }
.section-heading a { color:var(--accent); font-weight:600; text-decoration:none; font-size:14px; }
.section-note { color:var(--muted); font-size:14px; font-weight:600; }

/* ── Dropdown filter ─────────────────────────────────────────── */
.dropdown-browser {
  display:grid; grid-template-columns:repeat(4,minmax(0,1fr)) auto;
  gap:12px; align-items:end; padding:20px; margin-bottom:16px;
  border:1px solid var(--line); border-radius:var(--radius);
  background:var(--surface); box-shadow:var(--shadow);
}
.dropdown-browser label { display:grid; gap:6px; min-width:0; }
.dropdown-browser label span { color:var(--muted); font-size:12px; font-weight:700; text-transform:uppercase; letter-spacing:.05em; }
.dropdown-browser select {
  width:100%; height:42px; border:1px solid var(--line); border-radius:7px;
  padding:0 12px; background:var(--surface); color:var(--ink); font-size:14px;
  cursor:pointer; transition:border-color .15s;
}
.dropdown-browser select:focus { outline:none; border-color:var(--accent); }
.dropdown-browser button {
  height:42px; border:0; border-radius:7px; padding:0 16px;
  background:var(--ink); color:#fff; font-weight:700; font-size:14px; cursor:pointer;
  white-space:nowrap; transition:background .15s;
}
.dropdown-browser button:hover { background:#2d3f50; }
.dropdown-results { display:grid; grid-template-columns:repeat(3,minmax(0,1fr)); gap:10px; }
.load-more-btn {
  display:block; width:100%; margin-top:16px; padding:12px 0;
  border:2px dashed var(--line); border-radius:var(--radius);
  background:none; color:var(--muted); font-size:14px; font-weight:600;
  cursor:pointer; transition:all .15s;
}
.load-more-btn:hover { border-color:var(--accent); color:var(--accent); background:rgba(13,118,107,.04); }

/* ── Result items (search + dropdown) ───────────────────────── */
.search-results a,.dropdown-results a {
  display:flex; gap:12px; align-items:center;
  padding:12px 14px; border:1px solid var(--line); border-radius:var(--radius);
  background:var(--surface); text-decoration:none; transition:all .15s;
  box-shadow:var(--shadow);
}
.search-results a:hover,.dropdown-results a:hover { border-color:var(--accent); box-shadow:var(--shadow-md); transform:translateY(-1px); }
.search-results a strong,.dropdown-results a strong { display:block; font-size:15px; margin-bottom:3px; }
.search-results a span,.dropdown-results a span { color:var(--muted); font-size:13px; line-height:1.4; }
.result-thumb { width:56px; height:44px; object-fit:cover; border-radius:6px; flex-shrink:0; }

/* ── Facet grid ──────────────────────────────────────────────── */
.facet-grid { display:grid; grid-template-columns:repeat(auto-fill,minmax(160px,1fr)); gap:12px; }
.facet-card {
  display:flex; flex-direction:column; gap:6px; padding:18px;
  border:1px solid var(--line); border-radius:var(--radius); background:var(--surface);
  text-decoration:none; transition:all .15s; box-shadow:var(--shadow);
}
.facet-card:hover { border-color:var(--accent); box-shadow:var(--shadow-md); transform:translateY(-2px); }
.facet-card strong { font-size:15px; font-weight:700; }
.facet-card span { color:var(--muted); font-size:13px; }

/* ── Instrument card grid ────────────────────────────────────── */
.instrument-grid { display:grid; grid-template-columns:repeat(auto-fill,minmax(200px,1fr)); gap:16px; }
.instrument-card {
  display:flex; flex-direction:column;
  border:1px solid var(--line); border-radius:var(--radius); background:var(--surface);
  text-decoration:none; overflow:hidden; transition:all .15s;
  box-shadow:var(--shadow);
}
.instrument-card:hover { border-color:var(--accent); box-shadow:var(--shadow-md); transform:translateY(-2px); }
.card-thumb { display:block; width:100%; height:140px; object-fit:cover; flex-shrink:0; }
.card-thumb--empty { height:80px; background:linear-gradient(135deg,#f0f4f8,#e2e8f0); display:flex; align-items:center; justify-content:center; font-size:28px; color:#c5ced8; }
.card-body { padding:14px; display:flex; flex-direction:column; gap:4px; flex:1; }
.card-cat { color:var(--accent); font-size:11px; font-weight:700; text-transform:uppercase; letter-spacing:.05em; }
.card-title { font-size:16px; font-weight:700; line-height:1.3; }
.card-orig { color:var(--muted); font-size:12px; }
.card-meta { color:var(--muted); font-size:12px; margin-top:auto; }

/* ── Instrument detail page ──────────────────────────────────── */
.instrument-page { max-width:860px; }
.breadcrumb { display:flex; flex-wrap:wrap; gap:6px; color:var(--muted); font-size:13px; margin-bottom:24px; }
.breadcrumb a { text-decoration:none; color:var(--muted); }
.breadcrumb a:hover { color:var(--accent); }
.breadcrumb .sep { color:var(--line); }
.instrument-header { display:grid; gap:28px; margin-bottom:28px; align-items:start; }
.instrument-header.has-image { grid-template-columns:minmax(0,1fr) 240px; }
.header-text { display:flex; flex-direction:column; gap:4px; }
.header-image { position:sticky; top:80px; }
.original-name { color:var(--muted); font-size:15px; margin:6px 0 0; }
.soundscape-tag { color:var(--accent); font-size:14px; font-style:italic; margin:8px 0 0; line-height:1.5; }
.instrument-image { width:100%; border-radius:var(--radius); box-shadow:var(--shadow-md); object-fit:cover; }
.meta-grid {
  display:grid; grid-template-columns:repeat(auto-fill,minmax(175px,1fr));
  gap:10px; margin:0 0 32px; padding:0;
}
.meta-item { border:1px solid var(--line); border-radius:var(--radius); padding:14px; background:var(--surface); box-shadow:var(--shadow); }
.meta-item dt { color:var(--muted); font-size:12px; font-weight:600; text-transform:uppercase; letter-spacing:.05em; margin-bottom:6px; }
.meta-item dd { margin:0; font-weight:700; font-size:14px; line-height:1.4; word-break:break-word; }

/* ── YouTube embeds ──────────────────────────────────────────── */
.yt-grid { display:grid; grid-template-columns:repeat(2,minmax(0,1fr)); gap:16px; margin:16px 0 24px; }
.yt-embed { position:relative; padding-bottom:56.25%; border-radius:var(--radius); overflow:hidden; background:#000; box-shadow:var(--shadow-md); }
.yt-embed iframe { position:absolute; inset:0; width:100%; height:100%; border:0; }

/* ── Article body ────────────────────────────────────────────── */
.markdown-body { color:var(--ink2); line-height:1.8; font-size:16px; }
.markdown-body h2 { color:var(--ink); font-size:20px; margin:2em 0 .6em; padding-bottom:8px; border-bottom:1px solid var(--line); }
.markdown-body h3 { color:var(--ink); font-size:17px; margin:1.5em 0 .5em; }
.markdown-body p { margin:0 0 1.1em; }
.markdown-body ul,.markdown-body ol { margin:0 0 1em; padding-left:1.6em; }
.markdown-body li { margin-bottom:.35em; }
.markdown-body a { color:var(--blue); }
.markdown-body blockquote { margin:1em 0; padding:.5em 1em; border-left:3px solid var(--accent); color:var(--muted); background:var(--soft); border-radius:0 6px 6px 0; }
.markdown-body table { width:100%; border-collapse:collapse; margin-bottom:1.2em; font-size:15px; }
.markdown-body th { text-align:left; border-bottom:2px solid var(--line); padding:8px 12px; color:var(--muted); font-size:12px; text-transform:uppercase; }
.markdown-body td { border-bottom:1px solid var(--line); padding:9px 12px; }
.listen-button { display:inline-flex; align-items:center; min-height:42px; padding:0 18px; border-radius:7px; background:var(--accent); color:white; text-decoration:none; font-weight:700; font-size:14px; transition:background .15s; }
.listen-button:hover { background:var(--accent2); }
.source-note { color:var(--muted); font-size:14px; line-height:1.6; word-break:break-word; }
.source-note a { color:var(--blue); }

/* ── Tab bar ─────────────────────────────────────────────────── */
.tab-bar { display:flex; gap:4px; margin:0 0 20px; border-bottom:2px solid var(--line); padding-bottom:0; }
.tab-btn {
  padding:10px 20px; border:1px solid var(--line); border-bottom:none;
  border-radius:8px 8px 0 0; background:var(--soft); color:var(--muted);
  font-weight:700; font-size:14px; cursor:pointer; transition:all .15s;
  position:relative; top:2px;
}
.tab-btn:hover { color:var(--accent); }
.tab-btn.is-active { background:var(--surface); color:var(--accent); border-color:var(--line); border-bottom-color:var(--surface); }
.tab-pane { display:none; }
.tab-pane.is-active { display:block; }

/* ── Related instruments ─────────────────────────────────────── */
.related-section { margin-top:48px; padding-top:32px; border-top:1px solid var(--line); }
.related-heading { font-size:20px; margin-bottom:20px; color:var(--ink); }

/* ── Footer ──────────────────────────────────────────────────── */
.site-footer {
  border-top:1px solid var(--line); background:var(--surface);
  padding:20px 28px; margin-top:60px;
}
.footer-inner {
  max-width:1160px; margin:0 auto;
  display:flex; justify-content:space-between; align-items:center;
  flex-wrap:wrap; gap:12px; color:var(--muted); font-size:13px;
}
.footer-inner a { color:var(--blue); text-decoration:none; }
.visit-counter { font-size:12px; color:var(--muted); white-space:nowrap; }
.footer-nav { display:flex; gap:16px; }
.footer-nav a { color:var(--muted); text-decoration:none; }
.footer-nav a:hover { color:var(--accent); }

/* ── About page ─────────────────────────────────────────────── */
.about-hero {
  padding:48px 24px; text-align:center; color:#fff; border-radius:0 0 20px 20px;
  margin:-36px -24px 0; min-height:200px; display:flex; align-items:center; justify-content:center;
}
.about-hero-content { max-width:640px; }
.about-eyebrow { color:rgba(255,255,255,0.7); font-size:13px; font-weight:700; text-transform:uppercase; letter-spacing:.08em; margin:0 0 12px; }
.about-hero h1 { font-size:clamp(28px,4vw,42px); margin:0 0 16px; color:#fff; text-shadow:0 2px 8px rgba(0,0,0,0.3); }
.about-subtitle { font-size:18px; opacity:0.9; margin:0; }
.about-body { max-width:800px; margin:40px auto; }
.about-section { margin-bottom:40px; }
.about-section h2 { font-size:22px; margin:0 0 16px; padding-bottom:8px; border-bottom:2px solid var(--accent); display:inline-block; }
.about-text p { color:var(--ink2); line-height:1.9; font-size:16px; margin:0 0 16px; }
.about-author { display:flex; gap:24px; align-items:start; flex-wrap:wrap; }
.author-avatar { width:80px; height:80px; border-radius:50%; background:var(--accent); color:#fff; display:flex; align-items:center; justify-content:center; font-size:28px; font-weight:800; flex-shrink:0; }
.author-avatar img,img.author-avatar { width:80px; height:80px; border-radius:50%; object-fit:cover; display:block; }
.author-info { flex:1; min-width:200px; }
.author-info h3 { font-size:20px; margin:0 0 8px; }
.author-info p { color:var(--ink2); line-height:1.7; margin:0 0 8px; }
.about-grid { display:grid; grid-template-columns:1fr 1fr; gap:24px; }
.service-list { list-style:none; padding:0; margin:0; }
.service-list li { padding:10px 14px; border:1px solid var(--line); border-radius:8px; margin-bottom:8px; background:var(--surface); font-size:15px; display:flex; align-items:center; gap:10px; }
.service-icon { font-size:20px; }
.contact-info p { margin:0 0 12px; font-size:16px; }
.contact-info a { color:var(--blue); text-decoration:none; font-weight:600; }
.contact-info a:hover { text-decoration:underline; }

/* ── Feedback section ────────────────────────────────────────── */
.feedback-actions { margin-top:20px; display:flex; gap:12px; flex-wrap:wrap; }
.btn-line { display:inline-flex; align-items:center; gap:8px; padding:12px 24px; background:#06C755; color:#fff; border-radius:8px; text-decoration:none; font-weight:700; font-size:15px; transition:background .15s,transform .15s; box-shadow:0 2px 6px rgba(6,199,85,.3); }
.btn-line:hover { background:#05a648; transform:translateY(-1px); box-shadow:0 4px 12px rgba(6,199,85,.35); }
.btn-icon { font-size:20px; }

/* ── Image credit ───────────────────────────────────────────── */
.image-credit { font-size:12px; color:var(--muted); text-align:center; margin-top:6px; }

/* ── Badges ──────────────────────────────────────────────────── */
.badge-row { margin:8px 0 0; display:flex; gap:6px; }
.badge {
  display:inline-flex; padding:3px 8px; border-radius:4px;
  font-size:11px; font-weight:700;
}
.badge-hot { background:#fef2f0; color:#c2410c; border:1px solid #fed7c5; }
.badge-cold { background:#f0f5ff; color:#1e40af; border:1px solid #c5ddfd; }

/* ── Featured links ──────────────────────────────────────────── */
.featured-links { display:grid; grid-template-columns:repeat(3,minmax(0,1fr)); gap:14px; margin:0 0 40px; }
.featured-card {
  display:flex; flex-direction:column; gap:10px; border-radius:10px; padding:22px;
  text-decoration:none; transition:transform .15s,box-shadow .15s;
}
.featured-card:hover { transform:translateY(-2px); box-shadow:0 6px 20px rgba(0,0,0,.08); }
.featured-card.hot { background:linear-gradient(135deg,#fff5f0,#ffe8e0); border:1px solid #fdd4c5; }
.featured-card.cold { background:linear-gradient(135deg,#f0f7ff,#e0efff); border:1px solid #c5ddfd; }
.featured-card.random { background:linear-gradient(135deg,#f5f0ff,#ebe0ff); border:1px solid #d5c5fd; }
.featured-card strong { font-size:18px; }
.featured-card span { color:var(--muted); line-height:1.5; font-size:14px; }

/* ── Map section ─────────────────────────────────────────────── */
.world-map { width:100%; height:420px; border-radius:10px; border:1px solid var(--line); overflow:hidden; background:var(--soft); margin-bottom:8px; }
.map-hint { text-align:center; color:var(--muted); font-size:14px; margin:0 0 24px; }

/* ── Manager page ────────────────────────────────────────────── */
.manage-card {
  background:var(--surface); border:1px solid var(--line); border-radius:var(--radius);
  padding:28px; margin-bottom:24px; box-shadow:var(--shadow);
}
.manage-card h2 { font-size:18px; margin:0 0 8px; }
.manage-card p { color:var(--muted); font-size:14px; line-height:1.6; margin:0 0 18px; }
.btn {
  display:inline-flex; align-items:center; padding:10px 20px; border:0;
  border-radius:6px; font-size:15px; font-weight:700; cursor:pointer; text-decoration:none;
}
.btn-primary { background:var(--accent); color:#fff; }
.btn-primary:hover { background:var(--accent2); }
.upload-area { display:flex; gap:12px; align-items:center; flex-wrap:wrap; }
.upload-area input[type="file"] { flex:1; min-width:200px; padding:6px 0; font-size:14px; }
.upload-status { margin-top:12px; padding:8px 12px; border-radius:6px; font-size:14px; }
.upload-status.error { background:#fef2f0; color:#c2410c; border:1px solid #fed7c5; }
.upload-status.success { background:#f0fdf4; color:#15803d; border:1px solid #bbf7d0; }

/* ── Back to top ─────────────────────────────────────────────── */
.back-top {
  position:fixed; bottom:24px; right:24px;
  width:42px; height:42px; border-radius:50%;
  background:var(--accent); color:#fff; border:none;
  display:flex; align-items:center; justify-content:center;
  font-size:18px; cursor:pointer; opacity:0;
  transform:translateY(8px); transition:opacity .2s, transform .2s;
  z-index:50; box-shadow:0 2px 8px rgba(0,0,0,.2);
}
.back-top.visible { opacity:1; transform:none; }

/* ── Responsive ──────────────────────────────────────────────── */
@media (max-width:960px) {
  .instrument-grid { grid-template-columns:repeat(auto-fill,minmax(170px,1fr)); }
  .dropdown-browser { grid-template-columns:repeat(2,minmax(0,1fr)); }
  .dropdown-browser button { grid-column:1 / -1; }
  .dropdown-results { grid-template-columns:repeat(2,minmax(0,1fr)); }
  .stats { grid-template-columns:repeat(2,1fr); }
  .featured-links { grid-template-columns:repeat(2,minmax(0,1fr)); }
  .yt-grid { grid-template-columns:1fr !important; }
}
@media (max-width:700px) {
  .site-header { flex-direction:column; align-items:flex-start; gap:10px; padding:14px 18px; }
  .site-header nav { flex-wrap:wrap; gap:2px; }
  h1 { font-size:28px; }
  .instrument-grid,.stats,.meta-grid,.dropdown-browser,.dropdown-results,.featured-links { grid-template-columns:1fr !important; }
  .instrument-header,.instrument-header.has-image { grid-template-columns:1fr; }
  .header-image { position:static; }
  .page,.instrument-page { padding:20px 16px 60px; }
  .facet-grid { grid-template-columns:repeat(2,minmax(0,1fr)); }
  .world-map { height:280px; }
}

/* ── Sound journey page ────────────────────────────────────────── */
.sound-journey-page { max-width:780px; }
.journey-article { line-height:2; font-size:17px; color:var(--ink2); }
.journey-article h1 { font-size:clamp(26px,4vw,36px); margin-bottom:8px; }
.journey-article h2 { font-size:22px; margin:2em 0 .6em; padding-bottom:8px; border-bottom:1px solid var(--line); color:var(--ink); }
.journey-article h3 { font-size:18px; margin:1.8em 0 .4em; color:var(--ink); }
.journey-article p { margin:0 0 1.2em; }
.journey-images { display:flex; flex-wrap:wrap; gap:8px; justify-content:center; margin:1.2em 0; }
.journey-img { max-width:100%; height:auto; border-radius:6px; flex-shrink:0; object-fit:contain; }
.journey-images .journey-img { max-height:120px; width:auto; }
.experiment-box {
  margin:1.2em 0; padding:18px 20px;
  border:1px solid var(--accent); border-radius:8px;
  background:var(--surface); box-shadow:var(--shadow);
  font-size:15px; line-height:1.7;
}
.experiment-box p { margin:0; }
.experiment-box::before { content:"🧪 小實驗"; display:block; font-weight:700; color:var(--accent); font-size:13px; margin-bottom:8px; text-transform:uppercase; letter-spacing:.05em; }
.btn-back {
  display:inline-block; margin-top:16px; padding:10px 20px;
  background:var(--accent); color:#fff; border-radius:8px;
  text-decoration:none; font-weight:700;
}
.btn-back:hover { background:var(--accent2); }
@media (max-width:700px) {
  .journey-images .journey-img { max-height:80px; }
  .journey-article { font-size:16px; }
}
"""
    search_index = [
        {
            "title": item["title"],
            "original_name": item.get("original_name", ""),
            "category": item["category"],
            "country": item["country"],
            "era": item["era"],
            "sound_class": item.get("sound_class", ""),
            "url": site_url(f"/instruments/{item['slug']}/"),
            "image": safe_external_url(item.get("image", "")),
        }
        for item in instruments
    ]
    js = f"""
const SEARCH_INDEX = {json.dumps(search_index, ensure_ascii=False)};
const input = document.getElementById('site-search');
const results = document.getElementById('search-results');
const dropdownResults = document.getElementById('dropdown-results');
const dropdownCount = document.getElementById('dropdown-count');
const modeDropdown = document.getElementById('mode-dropdown');
const modeCards = document.getElementById('mode-cards');
const dropdownMode = document.getElementById('dropdown-mode');
const cardMode = document.getElementById('card-mode');
const filterControls = {{
  category: document.getElementById('filter-category'),
  country: document.getElementById('filter-country'),
  era: document.getElementById('filter-era'),
  sound_class: document.getElementById('filter-sound-class')
}};
const resetFilters = document.getElementById('filter-reset');

function appendResult(container, item) {{
  const link = document.createElement('a');
  link.href = item.url;
  if (item.image) {{
    const img = document.createElement('img');
    img.src = item.image;
    img.className = 'result-thumb';
    img.alt = '';
    img.loading = 'lazy';
    img.onerror = () => img.remove();
    link.append(img);
  }}
  const info = document.createElement('div');
  const title = document.createElement('strong');
  title.textContent = item.title;
  const meta = document.createElement('span');
  const parts = [item.category, item.country, item.era].filter(Boolean);
  meta.textContent = parts.join(' · ');
  info.append(title, meta);
  link.append(info);
  container.append(link);
}}

function setBrowseMode(mode) {{
  if (!dropdownMode || !cardMode) return;
  const useDropdown = mode !== 'cards';
  dropdownMode.hidden = !useDropdown;
  cardMode.hidden = useDropdown;
  if (modeDropdown) modeDropdown.classList.toggle('is-active', useDropdown);
  if (modeCards) modeCards.classList.toggle('is-active', !useDropdown);
  try {{ localStorage.setItem('wmi_browse_mode', mode); }} catch(e) {{}}
}}

modeDropdown?.addEventListener('click', () => setBrowseMode('dropdown'));
modeCards?.addEventListener('click', () => setBrowseMode('cards'));

// Restore saved mode
try {{
  const saved = localStorage.getItem('wmi_browse_mode');
  if (saved) setBrowseMode(saved); else setBrowseMode('dropdown');
}} catch(e) {{ setBrowseMode('dropdown'); }}

// Search box
if (input && results) {{
  input.addEventListener('input', () => {{
    const q = input.value.trim().toLowerCase();
    results.replaceChildren();
    if (!q) return;
    const scored = SEARCH_INDEX.flatMap(item => {{
      const t = item.title.toLowerCase();
      const o = (item.original_name || '').toLowerCase();
      const haystack = [t, o, item.category, item.country, item.era, item.sound_class].join(' ').toLowerCase();
      if (!haystack.includes(q)) return [];
      const score = t === q ? 0 : t.startsWith(q) ? 1 : o.startsWith(q) ? 2 : 3;
      return [{{item, score}}];
    }});
    scored.sort((a, b) => a.score - b.score);
    const hits = scored.slice(0, 20).map(x => x.item);
    for (const item of hits) appendResult(results, item);
  }});
}}

// Populate selects
function countValues(field) {{
  const counts = new Map();
  for (const item of SEARCH_INDEX) {{
    const v = item[field];
    if (!v) continue;
    counts.set(v, (counts.get(v) || 0) + 1);
  }}
  return [...counts.entries()].sort((a, b) => b[1] - a[1]);
}}

function fillSelect(select, field) {{
  if (!select) return;
  for (const [value, count] of countValues(field)) {{
    const option = document.createElement('option');
    option.value = value;
    option.textContent = `${{value}}（${{count}}）`;
    select.append(option);
  }}
}}

function selectedFilters() {{
  return Object.fromEntries(
    Object.entries(filterControls).map(([field, select]) => [field, select?.value || ''])
  );
}}

const PAGE_SIZE = 100;
let currentHits = [];
let shownCount = 0;

function renderNextPage() {{
  const batch = currentHits.slice(shownCount, shownCount + PAGE_SIZE);
  for (const item of batch) appendResult(dropdownResults, item);
  shownCount += batch.length;
  const oldBtn = document.getElementById('load-more-btn');
  if (oldBtn) oldBtn.remove();
  if (shownCount < currentHits.length) {{
    const btn = document.createElement('button');
    btn.id = 'load-more-btn';
    btn.className = 'load-more-btn';
    btn.textContent = `顯示更多（還有 ${{currentHits.length - shownCount}} 筆）`;
    btn.addEventListener('click', renderNextPage);
    dropdownResults.after(btn);
  }}
}}

function renderDropdownResults() {{
  if (!dropdownResults) return;
  dropdownResults.replaceChildren();
  const filters = selectedFilters();
  currentHits = SEARCH_INDEX.filter(item =>
    Object.entries(filters).every(([field, value]) => !value || item[field] === value)
  );
  shownCount = 0;
  if (dropdownCount) dropdownCount.textContent = `${{currentHits.length}} 筆`;
  renderNextPage();
  // Save filter state
  try {{ localStorage.setItem('wmi_filters', JSON.stringify(filters)); }} catch(e) {{}}
}}

if (dropdownResults) {{
  fillSelect(filterControls.category, 'category');
  fillSelect(filterControls.country, 'country');
  fillSelect(filterControls.era, 'era');
  fillSelect(filterControls.sound_class, 'sound_class');

  // Restore saved filters
  try {{
    const saved = JSON.parse(localStorage.getItem('wmi_filters') || '{{}}');
    for (const [field, value] of Object.entries(saved)) {{
      if (filterControls[field] && value) filterControls[field].value = value;
    }}
  }} catch(e) {{}}

  for (const select of Object.values(filterControls)) {{
    select?.addEventListener('change', renderDropdownResults);
  }}
  resetFilters?.addEventListener('click', () => {{
    for (const select of Object.values(filterControls)) {{
      if (select) select.value = '';
    }}
    try {{ localStorage.removeItem('wmi_filters'); }} catch(e) {{}}
    renderDropdownResults();
  }});
  renderDropdownResults();
}}

/* ── Back to top ──────────────────────────────────────────────── */
const backTop = document.getElementById('back-top');
if (backTop) {{
  window.addEventListener('scroll', () => {{
    backTop.classList.toggle('visible', window.scrollY > 400);
  }}, {{passive: true}});
  backTop.addEventListener('click', () => window.scrollTo({{top:0, behavior:'smooth'}}));
}}

/* ── Tab switching ──────────────────────────────────────────── */
document.querySelectorAll('.tab-btn').forEach(btn => {{
  btn.addEventListener('click', () => {{
    document.querySelectorAll('.tab-btn').forEach(b => b.classList.remove('is-active'));
    document.querySelectorAll('.tab-pane').forEach(p => p.classList.remove('is-active'));
    btn.classList.add('is-active');
    const pane = document.getElementById('tab-' + btn.dataset.tab);
    if (pane) pane.classList.add('is-active');
  }});
}});
"""
    write(OUTPUT_DIR / "assets" / "site.css", css.strip() + "\n")
    write(OUTPUT_DIR / "assets" / "search.js", js.strip() + "\n")
    write(OUTPUT_DIR / "search-index.json", json.dumps(search_index, ensure_ascii=False, indent=2))


def build_404(instruments):
    sample = random.Random(7).sample(instruments, min(8, len(instruments)))
    cards = "\n".join(card(item) for item in sample)
    body = f"""
    <main class="page">
      <section class="compact-hero">
        <p class="eyebrow">404</p>
        <h1>找不到頁面</h1>
        <p class="lead">這個頁面不存在，或已被移動。</p>
        <a href="{site_url('/')}" style="display:inline-block;margin-top:16px;padding:10px 20px;background:var(--accent);color:#fff;border-radius:8px;text-decoration:none;font-weight:700;">回首頁</a>
      </section>
      <section class="section">
        <div class="section-heading"><h2>隨機推薦</h2></div>
        <div class="instrument-grid">{cards}</div>
      </section>
    </main>"""
    write(OUTPUT_DIR / "404.html", page("頁面找不到", body))


def build_sitemap(instruments):
    site_base = f"{SITE_DOMAIN}{SITE_BASE_PATH}"
    today = datetime.date.today().isoformat()

    def u(path):
        return f"<url><loc>{site_base}{path}</loc><lastmod>{today}</lastmod></url>"

    # Collect all unique facet values from instruments
    categories = sorted(set(i["category"] for i in instruments if i.get("category")))
    countries = sorted(set(i["country"] for i in instruments if i.get("country")))
    eras = sorted(set(i["era"] for i in instruments if i.get("era")))
    sound_classes = sorted(set(i["sound_class"] for i in instruments if i.get("sound_class")))

    urls = []

    # Static top-level pages
    for path in ["/", "/instruments/", "/categories/", "/countries/", "/eras/",
                 "/sound-classes/", "/popular/", "/uncommon/", "/map/",
                 "/about/", "/theory/", "/vocal/", "/digitalmusic/",
                 "/sound-journey/", "/experience/", "/contact/"]:
        urls.append(u(path))

    # Instrument detail pages
    for item in instruments:
        urls.append(u(f"/instruments/{item['slug']}/"))

    # Category detail pages
    for name in categories:
        urls.append(u(f"/categories/{slugify(name)}/"))

    # Country detail pages
    for name in countries:
        urls.append(u(f"/countries/{slugify(name)}/"))

    # Era detail pages
    for name in eras:
        urls.append(u(f"/eras/{slugify(name)}/"))

    # Sound class detail pages
    for name in sound_classes:
        urls.append(u(f"/sound-classes/{slugify(name)}/"))

    # Theory detail pages (generated by build_vocal_extra.py)
    for i in range(6):  # 0–5
        urls.append(u(f"/theory/{i}/"))

    # Vocal chapter pages (generated by build_vocal_extra.py)
    for i in range(1, 51):  # 1–50
        urls.append(u(f"/vocal/{i}/"))

    # Digital music lesson pages (generated by build_vocal_extra.py)
    for i in range(1, 31):  # 1–30
        urls.append(u(f"/digitalmusic/{i}/"))

    # Sound journey pages
    urls.append(u("/sound-journey/all/"))
    for i in range(1, 7):  # 1–6
        urls.append(u(f"/sound-journey/{i}/"))

    xml = '<?xml version="1.0" encoding="UTF-8"?>\n<urlset xmlns="http://www.sitemaps.org/schemas/sitemap/0.9">\n'
    xml += "\n".join(urls) + "\n</urlset>\n"
    write(OUTPUT_DIR / "sitemap.xml", xml)


def build_special_pages(instruments):
    """Build popular and uncommon list pages."""
    popular = [i for i in instruments if i.get("is_popular")]
    uncommon = [i for i in instruments if i.get("is_uncommon")]

    if popular:
        page_dir = OUTPUT_DIR / "popular"
        page_dir.mkdir(parents=True, exist_ok=True)
        write(page_dir / "index.html", list_page("熱門樂器", popular, page_dir / "index.html", meta_description=f"熱門樂器 — {len(popular)} 件最受歡迎的世界樂器，包含鋼琴、吉他、小提琴、爵士鼓、卡林巴、手碟等知名樂器的介紹、聆賞與教學。世界聲音百科 by 隔壁織音人。"))

    if uncommon:
        page_dir = OUTPUT_DIR / "uncommon"
        page_dir.mkdir(parents=True, exist_ok=True)
        write(page_dir / "index.html", list_page("冷門樂器", uncommon, page_dir / "index.html", meta_description=f"冷門樂器 — {len(uncommon)} 件稀有珍奇的世界樂器，發掘來自世界各地的獨特聲音。世界聲音百科 by 隔壁織音人。"))

def build_map_page(instruments):
    """Build a standalone map page at /map/ with clickable markers."""
    from collections import defaultdict
    map_features = build_map_data(instruments)
    map_json = json.dumps(map_features, ensure_ascii=False)
    page_dir_ = OUTPUT_DIR / "map"
    page_dir_.mkdir(parents=True, exist_ok=True)

    slugs_json = json.dumps([i["slug"] for i in instruments], ensure_ascii=False)

    extra = '<link rel="stylesheet" href="https://unpkg.com/leaflet@1.9.4/dist/leaflet.css" crossorigin="">'
    body = f"""<main class="page">
  <section class="compact-hero">
    <p class="eyebrow">Geography</p>
    <h1>地圖導覽</h1>
    <p class="lead">點擊地圖上的標記，瀏覽該地區的所有樂器。</p>
  </section>
  <div id="world-map" class="world-map"></div>
  <p class="map-hint">每個圓點代表一個地區的樂器數量，點擊可查看該地區樂器列表。</p>
</main>
<script src="https://unpkg.com/leaflet@1.9.4/dist/leaflet.js" crossorigin=""></script>
<script>
(function() {{
  var mapData = {map_json};
  var container = document.getElementById('world-map');
  if (!container || !mapData.length) return;
  var map = L.map('world-map', {{ center: [20, 30], zoom: 2, minZoom: 1, maxZoom: 6, zoomControl: true, attributionControl: true }});
  L.tileLayer('https://{{s}}.tile.openstreetmap.org/{{z}}/{{x}}/{{y}}.png', {{ maxZoom: 18, attribution: '&copy; <a href="https://www.openstreetmap.org/copyright">OpenStreetMap</a>' }}).addTo(map);
  var bounds = [];
  mapData.forEach(function(item) {{
    var radius = Math.min(8 + item.count * 1.5, 30);
    var marker = L.circleMarker([item.lat, item.lng], {{ radius: radius, fillColor: '#0d766b', color: '#0a5c53', weight: 2, opacity: 1, fillOpacity: 0.6 }}).addTo(map);
    var tip = '<strong>' + item.name + '<\/strong><br>' + item.count + ' 件樂器';
    if (item.samples.length) tip += '<br><small>' + item.samples.join('、') + (item.count > 5 ? '…' : '') + '<\/small>';
    marker.bindTooltip(tip, {{ direction: 'top', offset: [0, -8] }});
    var links = '<div style="max-height:200px;overflow-y:auto">' + (item.urls || [item.url]).map(function(u) {{ return '<a href="' + u + '" style="display:block;padding:4px 0;color:#1d4ed8;text-decoration:none;border-bottom:1px solid #eee">' + u.split('/').filter(Boolean).pop() + '</a>'; }}).join('') + '</div>';
    marker.bindPopup('<strong>' + item.name + '</strong><br>' + item.count + ' 件樂器' + links, {{ maxWidth: 300 }});
    bounds.push([item.lat, item.lng]);
  }});
  if (bounds.length > 0) map.fitBounds(bounds, {{ padding: [30, 30], maxZoom: 4 }});
}})();
</script>"""
    write(page_dir_ / "index.html", page("地圖導覽", body, page_dir_ / "index.html", extra_head=extra, meta_description="世界樂器地圖導覽 — 透過互動式世界地圖探索各國樂器，點擊標記瀏覽該地區的所有樂器。世界聲音百科 by 隔壁織音人。"))


def build_manager_page(instruments):
    """Build a static manager page at /manage/ with Excel download and client-side upload."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    page_dir_ = OUTPUT_DIR / "manage"
    page_dir_.mkdir(parents=True, exist_ok=True)

    fm_fields = [
        ("title", "樂器名稱（繁體中文）"),
        ("original_name", "原文名稱（英文）"),
        ("category", "分類"),
        ("country", "來源地區"),
        ("era", "年代"),
        ("image", "圖片網址"),
        ("site_url", "網站連結"),
        ("sound_class", "發聲大類"),
        ("range", "音域"),
        ("instrument_key", "調性"),
        ("hs_class", "H-S 分類"),
        ("family", "家族"),
        ("playing_method", "演奏方式"),
        ("body_listening", "身體聆聽"),
        ("soundscape", "聲音景觀"),
        ("region_type", "區域類型"),
        ("youtube_ids", "YouTube ID"),
        ("introduction", "介紹內容"),
        ("history", "歷史背景"),
        ("timbre", "音色描述"),
        ("material", "樂器材質"),
        ("tutorial", "教學"),
    ]
    field_keys = [f[0] for f in fm_fields]

    # Generate Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "樂器總資料庫"
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="0F766E", end_color="0F766E", fill_type="solid")
    sub_font = Font(bold=True, color="666666", size=10)
    sub_fill = PatternFill(start_color="E6F4F1", end_color="E6F4F1", fill_type="solid")
    link_font = Font(color="1D4ED8", underline="single")

    for col_idx, (key, zh) in enumerate(fm_fields, 1):
        cell = ws.cell(row=1, column=col_idx, value=key)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell2 = ws.cell(row=2, column=col_idx, value=zh)
        cell2.font = sub_font
        cell2.fill = sub_fill
        cell2.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A3"

    # Parse body sections from markdown files for Excel export
    body_keys = {"introduction": "## \u4ecb\u7d39", "history": "## \u6b77\u53f2\u80cc\u666f",
                 "timbre": "## \u97f3\u8272\u63cf\u8ff0",
                 "listen": "## \u8046\u807d\u793a\u7bc4",
                 "representative": "## \u4ee3\u8868\u6027\u4f5c\u54c1",
                 "material": "## \u6a02\u5668\u6750\u8cea",
                 "tutorial": "## \u6559\u5b78"}

    md_files = sorted(OUTPUT_DIR.parent.parent.glob("content/instruments/*.md"))
    md_body_cache = {}
    for mf in md_files:
        slug = mf.stem
        md_text = mf.read_text(encoding="utf-8")
        # Split frontmatter from body
        fm_match = re.match(r"^---\s*\n.*?\n---", md_text, re.DOTALL)
        raw_body = md_text[fm_match.end():].strip() if fm_match else md_text.strip()
        sections = {}
        for key, heading in body_keys.items():
            pattern = re.compile(rf"^{re.escape(heading)}\s*\n(.*?)(?=\n## |\Z)", re.DOTALL | re.MULTILINE)
            m = pattern.search(raw_body)
            if m:
                sections[key] = m.group(1).strip()
        md_body_cache[slug] = sections

    for row_idx, item in enumerate(instruments, 3):
        slug = item.get("slug", "")
        body_sections = md_body_cache.get(slug, {})
        for col_idx, key in enumerate(field_keys, 1):
            val = item.get(key, "")
            # For body keys, get from parsed markdown
            if key in ("introduction", "history", "timbre", "material", "tutorial"):
                val = body_sections.get(key, "")
            if val:
                cell = ws.cell(row=row_idx, column=col_idx, value=str(val))
                if key in ("site_url", "image") and str(val).startswith("http"):
                    cell.font = link_font
                    cell.hyperlink = str(val)

    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 50)

    # Add theory basics sheet
    ws_theory = wb.create_sheet(title="樂理基礎")
    theory_fields = [("topic", "主題"), ("content", "內容")]
    for col_idx, (key, zh) in enumerate(theory_fields, 1):
        cell = ws_theory.cell(row=1, column=col_idx, value=key)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell2 = ws_theory.cell(row=2, column=col_idx, value=zh)
        cell2.font = sub_font
        cell2.fill = sub_fill
        cell2.alignment = Alignment(horizontal="center")
    for row_idx, tab in enumerate(get_theory_data(), 3):
        ws_theory.cell(row=row_idx, column=1, value=tab["label"]).font = Font(bold=True, size=11)
        plain_text = re.sub(r"<[^>]+>", "", tab["content"]).strip()
        cell = ws_theory.cell(row=row_idx, column=2, value=plain_text)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws_theory.column_dimensions["A"].width = 16
    ws_theory.column_dimensions["B"].width = 120
    ws_theory.freeze_panes = "A3"

    excel_path = OUTPUT_DIR / "assets" / "instruments_database.xlsx"
    excel_path.parent.mkdir(parents=True, exist_ok=True)
    try:
        wb.save(str(excel_path))
    except PermissionError:
        print("Warning: Excel file locked, skipping")

    body = """<main class="page" style="max-width:720px;">
  <section class="compact-hero">
    <p class="eyebrow">Management</p>
    <h1>管理者頁面</h1>
  </section>

  <div id="manage-app">
    <div class="manage-card">
      <h2>下載 Excel 樂器總資料庫</h2>
      <p>將所有樂器的 Markdown 檔案匯出為 Excel 格式。包含所有 frontmatter 欄位與介紹、歷史、音色描述等內容。第一列為英文欄位名稱，第二列為中文說明。</p>
      <a class="btn btn-primary" href="../assets/instruments_database.xlsx" download>下載 Excel</a>
    </div>

    <div class="manage-card">
      <h2>上傳 Excel 還原 Markdown</h2>
      <p>上傳一個 Excel 檔案（格式需與下載版本相同），系統會在瀏覽器中自動處理每一列資料，還原為獨立的 Markdown 檔案，並打包成 ZIP 壓縮檔供下載。</p>
      <div class="upload-area">
        <input type="file" id="excel-upload" accept=".xlsx">
        <button class="btn btn-primary" id="process-excel">上傳並下載 ZIP</button>
      </div>
      <div id="upload-status" class="upload-status"></div>
    </div>
  </div>


</main>

<script src="https://cdnjs.cloudflare.com/ajax/libs/xlsx/0.18.5/xlsx.full.min.js"></script>
<script src="https://cdnjs.cloudflare.com/ajax/libs/jszip/3.10.1/jszip.min.js"></script>
<script>
(function() {
  if (sessionStorage.getItem('manage_pass') === 'ok') {
    document.getElementById('manage-locked').style.display = 'none';
    document.getElementById('manage-app').style.display = 'block';
  } else {
    document.getElementById('manage-locked').style.display = 'block';
    document.getElementById('manage-app').style.display = 'none';
    setTimeout(function() {
      var c = prompt('管理者密碼：');
      if (c && c === atob('NTIwMTMxNA==')) {
        sessionStorage.setItem('manage_pass', 'ok');
        document.getElementById('manage-locked').style.display = 'none';
        document.getElementById('manage-app').style.display = 'block';
      } else if (c !== null) {
        document.getElementById('password-error').textContent = '密碼錯誤';
      }
    }, 200);
  }

  function slugify(name) {
    if (!name) return 'instrument';
    return name.toLowerCase().trim().replace(/[^a-z0-9]+/g, '-').replace(/-+/g, '-').replace(/^-|-$/g, '') || 'instrument';
  }

  document.getElementById('process-excel')?.addEventListener('click', function() {
    if (typeof XLSX === 'undefined' || typeof JSZip === 'undefined') {
      document.getElementById('upload-status').textContent = '函式庫載入中，請稍後再試...';
      document.getElementById('upload-status').className = 'upload-status error';
      return;
    }
    var fileInput = document.getElementById('excel-upload');
    var status = document.getElementById('upload-status');
    if (!fileInput || !fileInput.files.length) {
      status.textContent = '請選擇一個 Excel 檔案';
      status.className = 'upload-status error';
      return;
    }
    var file = fileInput.files[0];
    var reader = new FileReader();
    reader.onload = function(e) {
      try {
        var data = new Uint8Array(e.target.result);
        var workbook = XLSX.read(data, {type: 'array'});
        var sheet = workbook.Sheets[workbook.SheetNames[0]];
        var rows = XLSX.utils.sheet_to_json(sheet, {header: 1});
        if (rows.length < 3) {
          status.textContent = '檔案格式錯誤：缺少足夠的資料列（需要標題列 + 資料列）';
          status.className = 'upload-status error';
          return;
        }
        var headers = rows[0] || [];
        var zip = new JSZip();
        var count = 0;
        for (var i = 2; i < rows.length; i++) {
          var row = rows[i];
          if (!row || !row[0]) continue;
          var fmLines = ['---'];
          var sections = {};
          var origName = '';
          for (var j = 0; j < headers.length; j++) {
            var key = String(headers[j] || '').trim();
            var val = row[j] !== undefined ? String(row[j]).trim() : '';
            if (!key || !val) continue;
            if (key === 'original_name') origName = val;
            if (key === 'introduction' || key === 'history' || key === 'timbre') {
              sections[key] = val;
            } else {
              fmLines.push(key + ': ' + val);
            }
          }
          fmLines.push('---');
          var bodyParts = [];
          if (sections.introduction) bodyParts.push('\n## 介紹\n\n' + sections.introduction);
          if (sections.history) bodyParts.push('\n## 歷史背景\n\n' + sections.history);
          if (sections.timbre) bodyParts.push('\n## 音色描述\n\n' + sections.timbre);
          var fullContent = fmLines.join('\n') + bodyParts.join('\n') + '\n';
          var filename = slugify(origName || ('instrument_' + (i + 1))) + '.md';
          zip.file(filename, fullContent);
          count++;
        }
        if (count === 0) {
          status.textContent = '找不到有效的樂器資料，請確認 Excel 格式是否正確。';
          status.className = 'upload-status error';
          return;
        }
        zip.generateAsync({type: 'blob'}).then(function(blob) {
          var url = URL.createObjectURL(blob);
          var a = document.createElement('a');
          a.href = url;
          a.download = 'instruments_markdown.zip';
          document.body.appendChild(a);
          a.click();
          document.body.removeChild(a);
          setTimeout(function() { URL.revokeObjectURL(url); }, 10000);
          status.textContent = '成功生成 ' + count + ' 個 .md 檔案，已開始下載。';
          status.className = 'upload-status success';
        });
      } catch(err) {
        status.textContent = '處理失敗：' + err.message;
        status.className = 'upload-status error';
        console.error(err);
      }
    };
    reader.readAsArrayBuffer(file);
  });
})();
</script>"""
    write(page_dir_ / "index.html", page("管理者頁面", body, page_dir_ / "index.html", meta_extra='<meta name="robots" content="noindex">'))



def build_about_page():
    """Build the About page with website info and author details."""
    page_dir_ = OUTPUT_DIR / "about"
    page_dir_.mkdir(parents=True, exist_ok=True)
    bg_url = "https://yt3.googleusercontent.com/6nBZ7RVoXGMH2fuMPWiju_tpAET9D-qVkOhg1HjGqh8m9EaO-u9wO_oHVA12Sy0DzoKn7mGVmA=w1707-fcrop64=1,00005a57ffffa5a8-k-c0xffffffff-no-nd-rj"
    # Copy author logo into output assets
    src_logo = BASE_DIR / "static" / "author-logo.jpg"
    dst_logo = OUTPUT_DIR / "assets" / "author-logo.jpg"
    dst_logo.parent.mkdir(parents=True, exist_ok=True)
    if src_logo.exists():
        import shutil as _shutil
        _shutil.copy2(str(src_logo), str(dst_logo))
    logo_url = resolve_url(page_dir_ / "index.html", "/assets/author-logo.jpg")
    body = f"""<main class="about-page">
  <section class="about-hero" style="background:linear-gradient(135deg,rgba(0,0,0,0.7),rgba(0,0,0,0.4)),url({bg_url}) center/cover no-repeat;">
    <div class="about-hero-content">
      <p class="about-eyebrow">World Musical Instruments Encyclopedia</p>
      <h1>關於世界聲音百科</h1>
      <p class="about-subtitle">循著聲音，走進不同文化的現場</p>
    </div>
  </section>

  <div class="about-body">
    <section class="about-section">
      <h2>創建網站理念</h2>
      <div class="about-text">
        <p>這個網站的初衷，源於我們對世界樂器長期的熱愛與好奇。每一件樂器，不只是發聲的工具，更是地方文化、工藝技術、生活信仰與情感記憶的載體。當我們聽見陌生的音色，有時感到遙遠，有時卻產生莫名的共鳴。這正是世界樂器最迷人的地方。</p>

        <h3 style="font-size:18px;margin:1.5em 0 .5em;color:var(--ink);">為什麼我們需要這個平台？</h3>
        <p>在現今的資訊環境下，關於樂器的知識多半分散在學術論文、零散文章或影音平台中。對於探索者而言，很難從樂器分類、地域分佈、構造、演奏方式到文化脈絡之間建立完整的理解。</p>
        <p>因此，我們致力於建立一個兼具知識性、可讀性與探索感的百科平台。我們將世界各地的樂器整理為系統化的知識體系，讓讀者能像翻閱地圖般，循著聲音的軌跡，走進不同文化的現場。</p>

        <h3 style="font-size:18px;margin:1.5em 0 .5em;color:var(--ink);">AI 輔助與知識查證</h3>
        <p>我們深知知識的準確性是百科的核心。網站內容由我們的核心團隊整理，並結合人工智慧（AI）技術進行交叉比對與查證，確保資料的深度與脈絡的一致性。</p>
        <p>儘管我們竭力追求精確，但在龐大的知識建置過程中，難免會有遺漏或錯誤。若您在閱讀過程中發現任何資訊有誤，或有更精準的見解，非常歡迎透過網站的回饋機制告訴我們。您的參與，是讓這份知識地圖更完整的關鍵。</p>

        <h3 style="font-size:18px;margin:1.5em 0 .5em;color:var(--ink);">三位一體的聲音知識地圖</h3>
        <p>隨著發展，我們意識到音樂的本質是由三個面向構成的：</p>
        <ul>
          <li><strong>世界樂器（聲音的載體）</strong>：從構造特色、演奏方式到文化背景，讓樂器不再是冷冰冰的條目，而是激發創作靈感的門戶。</li>
          <li><strong>人聲與歌唱（聲音的源頭）</strong>：每個人都擁有一副與生俱來的樂器。我們開設了從發聲啟蒙、身體覺察到混聲技術與聲學研究的教學體系，助您找回自在歌唱的自信。</li>
          <li><strong>錄音後製（聲音的塑造）</strong>：無論是宅錄 DEMO、Podcast 收音，還是進階的混音與母帶處理，我們規劃了從器材指南到製作觀念的實戰指南，幫助您將作品打磨得更臻完美。</li>
        </ul>
        <p>從樂器、人聲到後製，我們將這三個面向整合在同一個平台上，讓音樂的學習不再是碎片化的，而是一幅互相連結的完整知識圖譜。無論您是音樂創作者、教育工作者、學生，還是對聲音感到好奇的探索者，這裡都有適合您的章節。</p>
        <p style="margin-top:20px;padding:14px 18px;background:var(--accent);color:#fff;border-radius:8px;display:inline-block;font-weight:700;">創站日期：2026 年 6 月 25 日</p>
      </div>
    </section>

    <section class="about-section">
      <h2>作者</h2>
      <div class="about-author">
        <img class="author-avatar" src="{logo_url}" alt="隔壁織音人" loading="lazy">
        <div class="author-info">
          <h3>隔壁織音人</h3>
          <p>為隔壁音樂工作室 Next Door Music 的音樂團體</p>
          <p>隔壁音樂匯聚熱愛音樂夥伴的創意單位。我們將每位合作夥伴視為珍貴的「音樂鄰居」，在此紀錄彼此的創作火花。</p>
          <p>我們致力探索音符的無限可能，打造高品質聽覺體驗，並提供專業音樂製作服務。</p>
        </div>
      </div>
    </section>

    <div class="about-grid">
      <section class="about-section">
        <h2>服務項目 Service</h2>
        <ul class="service-list">
          <li><span class="service-icon">🎵</span>編曲製作｜翻唱改編．風格重塑</li>
          <li><span class="service-icon">🎼</span>詞曲訂製｜原創詞曲</li>
          <li><span class="service-icon">🎚️</span>後期處理｜人聲修音．混音</li>
          <li><span class="service-icon">🎤</span>人聲錄製｜DEMO 代唱．導唱</li>
        </ul>
      </section>

    </div>
  </div>
</main>"""
    write(page_dir_ / "index.html", page("關於", body, page_dir_ / "index.html", meta_description="世界聲音百科的創建理念、作者介紹與服務項目。收錄世界樂器、人聲歌唱教學、音樂理論與錄音製作知識。"))

def get_theory_data():
    """Stub - theory data now read from content/musictheory/*.md by build_vocal_extra.py"""
    return []


def build_theory_page():
    """Stub - theory page now generated by build_vocal_extra.py"""
    pass

def build_robots(instruments):
    """Generate robots.txt allowing all crawlers and referencing sitemap."""
    sitemap_url = f"{SITE_DOMAIN}{SITE_BASE_PATH}/sitemap.xml"
    robots_txt = "User-agent: *\nAllow: /\nSitemap: " + sitemap_url + "\n"
    write(OUTPUT_DIR / "robots.txt", robots_txt)


def build_sound_journey_pages():
    """Build the sound-journey section: index (旅圖門口), journey placeholders, and all articles page."""
    import re as _re
    sj_content_dir = BASE_DIR / "content" / "sound-journey"
    sj_output_dir = OUTPUT_DIR / "sound-journey"

    if not sj_content_dir.exists():
        return

    # --- Build index page (旅圖門口) ---
    index_md = sj_content_dir / "index.md"
    if index_md.exists():
        meta, body = parse_frontmatter(index_md.read_text(encoding="utf-8"))

        # Pre-process markdown images to HTML so they work inside div blocks
        body = _re.sub(
            r'!\[([^\]]*)\]\(([^)]+)\)',
            r'<img src="\2" alt="\1" class="journey-img" loading="lazy">',
            body,
        )

        body_html = markdown.markdown(body, extensions=["extra"], output_format="html5")

        # Copy images to output
        img_src = sj_content_dir / "images"
        img_dst = sj_output_dir / "images"
        if img_src.exists():
            if img_dst.exists():
                shutil.rmtree(str(img_dst))
            shutil.copytree(str(img_src), str(img_dst))

        page_body = f"""
        <main class="page sound-journey-page">
          <article class="journey-article">
            {body_html}
          </article>
        </main>
        """
        write(
            sj_output_dir / "index.html",
            page(
                meta.get("title", "旅圖門口"),
                page_body,
                sj_output_dir / "index.html",
                meta_description="隔壁家的世界聲音旅圖：從古老樂器到現代耳朵的聲音巡禮 — 循著聲音，走進不同文化的現場。",
            ),
        )

    # --- Build journey placeholder pages (旅圖一 ~ 旅圖六) ---
    journeys = [
        ("1", "旅圖一｜氣息離開身體", "氣息離開身體"),
        ("2", "旅圖二｜懷裡與大地的弦", "懷裡與大地的弦"),
        ("3", "旅圖三｜天空裡拉長的弦", "天空裡拉長的弦"),
        ("4", "旅圖四｜地心與手邊的微光", "地心與手邊的微光"),
        ("5", "旅圖五｜城市與星塵的節奏", "城市與星塵的節奏"),
        ("6", "旅圖六｜按鍵打開房間，眾聲走向廣場", "按鍵打開房間，眾聲走向廣場"),
    ]
    for num, title, subtitle in journeys:
        page_dir = sj_output_dir / num
        body = f"""
        <main class="page">
          <section class="compact-hero">
            <p class="eyebrow">聲音旅圖</p>
            <h1>{escape(title)}</h1>
            <p class="lead" style="font-size:1.2em;margin:1em 0;">📝 此篇章撰寫中，敬請期待</p>
            <p style="color:var(--muted);">{escape(subtitle)}</p>
            <a href="../" class="btn-back">← 回到旅圖門口</a>
          </section>
        </main>
        """
        write(
            page_dir / "index.html",
            page(
                title,
                body,
                page_dir / "index.html",
                meta_description=f"隔壁家的世界聲音旅圖 — {title}。篇章撰寫中，敬請期待。",
            ),
        )

    # --- Build "全部文章" page ---
    all_items = [
        ("../", "旅圖門口", "耳朵，可以醒來了", ""),
        ("../1/", "旅圖一｜氣息離開身體", "氣息離開身體", "更新中"),
        ("../2/", "旅圖二｜懷裡與大地的弦", "懷裡與大地的弦", "更新中"),
        ("../3/", "旅圖三｜天空裡拉長的弦", "天空裡拉長的弦", "更新中"),
        ("../4/", "旅圖四｜地心與手邊的微光", "地心與手邊的微光", "更新中"),
        ("../5/", "旅圖五｜城市與星塵的節奏", "城市與星塵的節奏", "更新中"),
        ("../6/", "旅圖六｜按鍵打開房間，眾聲走向廣場", "按鍵打開房間，眾聲走向廣場", "更新中"),
    ]
    items_html = ""
    for url, title, subtitle, status in all_items:
        status_badge = f'<span class="badge badge-cold">{escape(status)}</span>' if status else ""
        items_html += f'<a class="facet-card" href="{url}"><strong>{escape(title)}</strong><span>{escape(subtitle)}</span>{status_badge}</a>'

    body = f"""
    <main class="page">
      <section class="compact-hero">
        <p class="eyebrow">Sound Journey</p>
        <h1>全部文章</h1>
        <p class="lead">隔壁家的世界聲音旅圖 — 全部篇章一覽</p>
      </section>
      <div class="facet-grid">{items_html}</div>
    </main>
    """
    write(
        sj_output_dir / "all" / "index.html",
        page(
            "全部文章",
            body,
            meta_description="隔壁家的世界聲音旅圖全部篇章一覽 — 從氣息、弦、鼓到電子聲音，六段旅圖六種靠近世界的方式。",
        ),
    )
def main():
    global _TOTAL_INSTRUMENTS
    instruments = read_instruments()
    _TOTAL_INSTRUMENTS = len(instruments)
    if OUTPUT_DIR.exists():
        def on_rm_error(func, path, exc_info):
            pass
        shutil.rmtree(OUTPUT_DIR, onerror=on_rm_error)
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    build_assets(instruments)
    build_index(instruments)
    build_map_page(instruments)
    build_about_page()
    build_sound_journey_pages()
    build_theory_page()
    build_manager_page(instruments)
    build_detail_pages(instruments)
    build_special_pages(instruments)
    build_instruments_list_page(instruments)
    build_facet_pages(instruments, "category", "categories", "分類")
    build_facet_pages(instruments, "sound_class", "sound-classes", "發聲分類")
    build_facet_pages(instruments, "country", "countries", "國家/地區")
    build_facet_pages(instruments, "era", "eras", "年代")
    build_404(instruments)
    build_sitemap(instruments)
    build_robots(instruments)
    # Copy sitemap and robots.txt to project root for local development
    shutil.copy2(OUTPUT_DIR / "sitemap.xml", BASE_DIR / "sitemap.xml")
    shutil.copy2(OUTPUT_DIR / "robots.txt", BASE_DIR / "robots.txt")
    # Sync assets to project root (tracked by git for direct access)
    root_assets = BASE_DIR / "assets"
    root_assets.mkdir(parents=True, exist_ok=True)
    out_assets = OUTPUT_DIR / "assets"
    if out_assets.exists():
        for fname in ["instruments_database.xlsx", "site.css", "search.js",
                       "random-instrument.js", "map-data.json", "author-logo.jpg"]:
            src = out_assets / fname
            if src.exists():
                shutil.copy2(str(src), str(root_assets / fname))
    write(OUTPUT_DIR / ".nojekyll", "")
    # Write ads.txt for Google AdSense
    write(OUTPUT_DIR / "ads.txt", "google.com, pub-6561686484716387, DIRECT, f08c47fec0942fa0\n")
    # Copy Google Search Console verification file to output root
    gsc_file = BASE_DIR / "googled81b331c3a5b66a0.html"
    if gsc_file.exists():
        shutil.copy2(gsc_file, OUTPUT_DIR / "googled81b331c3a5b66a0.html")
    print(f"Built {len(instruments)} instruments into {OUTPUT_DIR}")


if __name__ == "__main__":
    main()
