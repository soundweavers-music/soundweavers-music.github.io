"""
遷移 content/instruments/*.md 的 YAML frontmatter 到新版欄位結構

使用 Excel 的「01_單一樂器主表_v1_0」作為新欄位資料來源，
搭配「05_欄位對照_v1_0」的映射規則進行轉換。
"""

import os
import re
import sys
import openpyxl

EXCEL_PATH = r'C:\Users\timmychi\Downloads\world_instrument_database_restructure_v1_2.xlsx'
INSTRUMENTS_DIR = r'D:\nextdoor\python-python-django-world-musical-instrument\content\instruments'
SHEET_INDEX = 1  # 01_單一樂器主表_v1_0

# ─── 從 Excel 讀取新欄位資料 ──────────────────────────────
print('正在讀取 Excel 01_單一樂器主表_v1_0 ...')
wb = openpyxl.load_workbook(EXCEL_PATH)
ws = wb[wb.sheetnames[SHEET_INDEX]]

raw_headers = [str(c.value).strip() if c.value else '' for c in ws[1]]

# Build lookups
slug_lookup = {}       # by source_url slug
name_lookup = {}       # by title_original (lowercase)
zh_lookup = {}         # by title_zh
records_without_url = []  # entries with no source_url

for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column, values_only=True), start=2):
    row_data = {}
    for i, val in enumerate(row):
        if i < len(raw_headers) and raw_headers[i]:
            row_data[raw_headers[i]] = str(val).strip() if val is not None else ''

    title_zh = row_data.get('title_zh（中文名）', '')
    title_orig = row_data.get('title_original（英文／原文名）', '')
    source_url = row_data.get('source_url（來源網址）', '')

    if not title_zh and not title_orig:
        continue  # skip empty rows

    m = re.search(r'/instruments/([^/]+)/', source_url)
    if m:
        slug = m.group(1).lower()
        slug_lookup[slug] = row_data
    else:
        records_without_url.append((title_zh, title_orig, row_data))

    if title_orig:
        key = title_orig.strip().lower()
        key = key.split('/')[0].strip()
        name_lookup[key] = row_data

    if title_zh:
        zh_lookup[title_zh.strip()] = row_data

print(f'  slug_lookup: {len(slug_lookup)} entries')
print(f'  name_lookup: {len(name_lookup)} entries')
print(f'  zh_lookup: {len(zh_lookup)} entries')
print(f'  records_without_url: {len(records_without_url)} entries')

# ─── 新欄位順序（frontmatter 輸出順序）───────────────────
NEW_FIELD_ORDER = [
    'class_code（主分類代碼）',
    'frontend_class（前台主分類）',
    'subcategory（子分類）',
    'title_zh（中文名）',
    'title_original（英文／原文名）',
    'family_std（樂器家族）',
    'sound_hs（發聲原理／H-S）',
    'playing_method（演奏方式）',
    'interface_tags（操作介面）',
    'region_culture（地域／文化）',
    'listening_sound_tags（聆聽與聲音標籤）',
    'ensemble_links（常見合奏／編制）',
    'verification_status（查證狀態）',
    'issue_note（核實／修正備註）',
    'source_url（來源網址）',
]

# ─── 處理每個檔案 ────────────────────────────────────────
md_files = sorted([f for f in os.listdir(INSTRUMENTS_DIR) if f.endswith('.md')])

matched_count = 0
unmatched_count = 0
unmatched_files = []

print(f'\n開始處理 {len(md_files)} 個檔案 ...')

for filename in md_files:
    filepath = os.path.join(INSTRUMENTS_DIR, filename)
    slug = filename.replace('.md', '').lower()

    with open(filepath, 'r', encoding='utf-8') as f:
        content = f.read()

    fm_match = re.match(r'^---\n(.*?)\n---', content, re.DOTALL)
    if not fm_match:
        print(f'  WARNING: 無法解析 frontmatter: {filename}')
        continue

    old_fm_text = fm_match.group(1)
    body = content[fm_match.end():].strip()

    # Parse old frontmatter
    old_fm = {}
    for line in old_fm_text.split('\n'):
        m2 = re.match(r'^(\w[\w_-]*)\s*:\s*(.*?)$', line)
        if m2:
            old_fm[m2.group(1).strip()] = m2.group(2).strip()

    # ─── 尋找符合的 Excel 記錄 ───
    row_data = slug_lookup.get(slug)

    if row_data is None:
        orig_name = old_fm.get('original_name', '').strip().lower()
        orig_name = orig_name.split('/')[0].strip()
        row_data = name_lookup.get(orig_name)

    if row_data is None:
        title_val = old_fm.get('title', '').strip()
        row_data = zh_lookup.get(title_val)

    if row_data is None:
        title_val = old_fm.get('title', '').strip()
        orig_name = old_fm.get('original_name', '').strip().lower()
        for tzh, torig, rec in records_without_url:
            if tzh == title_val:
                row_data = rec
                break
            if torig and torig.strip().lower().split('/')[0].strip() == orig_name:
                row_data = rec
                break

    # ─── 建構新 frontmatter ───
    new_fm_lines = []

    if row_data:
        # 從 Excel 記錄取得新欄位值
        for field_key in NEW_FIELD_ORDER:
            val = row_data.get(field_key, '')
            short_key = field_key.split('（')[0]
            if val:
                if '\n' in val:
                    new_fm_lines.append(f'{short_key}: >')
                    for line in val.split('\n'):
                        trimmed = line.strip()
                        if trimmed:
                            new_fm_lines.append(f'  {trimmed}')
                else:
                    new_fm_lines.append(f'{short_key}: {val}')
            else:
                new_fm_lines.append(f'{short_key}: ')
        matched_count += 1
    else:
        # ─── 無 Excel 記錄 → 舊欄位最佳映射 ───
        unmatched_count += 1
        unmatched_files.append(filename)

        # 直接映射的欄位
        field_mapping = {
            'title': 'title_zh',
            'original_name': 'title_original',
            'site_url': 'source_url',
            'playing_method': 'playing_method',
        }
        used_old_keys = set()

        for old_key, new_key in field_mapping.items():
            if old_key in old_fm:
                new_fm_lines.append(f'{new_key}: {old_fm[old_key]}')
                used_old_keys.add(old_key)

        # 新欄位（無舊對應，留空）
        new_fm_lines.append(f'class_code: ')
        new_fm_lines.append(f'frontend_class: ')
        new_fm_lines.append(f'subcategory: ')

        # family → family_std
        if 'family' in old_fm:
            new_fm_lines.append(f'family_std: {old_fm["family"]}')
            used_old_keys.add('family')
        else:
            new_fm_lines.append(f'family_std: ')

        # sound_class + hs_class → sound_hs
        sound_parts = []
        if 'sound_class' in old_fm:
            sound_parts.append(old_fm['sound_class'])
            used_old_keys.add('sound_class')
        if 'hs_class' in old_fm:
            sound_parts.append(old_fm['hs_class'])
            used_old_keys.add('hs_class')
        if sound_parts:
            new_fm_lines.append(f'sound_hs: {"｜".join(sound_parts)}')
        else:
            new_fm_lines.append(f'sound_hs: ')

        new_fm_lines.append(f'interface_tags: ')

        # country + region_type → region_culture
        region_parts = []
        if 'country' in old_fm:
            region_parts.append(old_fm['country'])
            used_old_keys.add('country')
        if 'region_type' in old_fm:
            region_parts.append(old_fm['region_type'])
            used_old_keys.add('region_type')
        if region_parts:
            new_fm_lines.append(f'region_culture: {"／".join(region_parts)}')
        else:
            new_fm_lines.append(f'region_culture: ')

        # body_listening + soundscape → listening_sound_tags
        listen_parts = []
        if 'body_listening' in old_fm:
            listen_parts.append(old_fm['body_listening'])
            used_old_keys.add('body_listening')
        if 'soundscape' in old_fm:
            listen_parts.append(old_fm['soundscape'])
            used_old_keys.add('soundscape')
        if listen_parts:
            new_fm_lines.append(f'listening_sound_tags: {"｜".join(listen_parts)}')
        else:
            new_fm_lines.append(f'listening_sound_tags: ')

        new_fm_lines.append(f'ensemble_links: ')
        new_fm_lines.append(f'verification_status: ')
        new_fm_lines.append(f'issue_note: ')

        # 確保 title_zh 和 title_original 有在正確位置
        has_zh = any(l.startswith('title_zh:') for l in new_fm_lines)
        has_orig = any(l.startswith('title_original:') for l in new_fm_lines)
        if not has_zh and 'title' in old_fm:
            new_fm_lines.insert(0, f'title_zh: {old_fm["title"]}')
        if not has_orig and 'original_name' in old_fm:
            for i, l in enumerate(new_fm_lines):
                if l.startswith('title_zh:'):
                    new_fm_lines.insert(i+1, f'title_original: {old_fm["original_name"]}')
                    break

    # ─── 寫入檔案 ───
    new_content = '---\n' + '\n'.join(new_fm_lines) + '\n---\n\n' + body + '\n'

    with open(filepath, 'w', encoding='utf-8') as f:
        f.write(new_content)

    if (matched_count + unmatched_count) % 200 == 0:
        print(f'  已處理 {matched_count + unmatched_count}/{len(md_files)} ...')

# ─── 輸出統計 ────────────────────────────────────────────
print(f'\n✅ 完成!')
print(f'  成功匹配更新: {matched_count} 個檔案')
print(f'  最佳映射（無記錄）: {unmatched_count} 個檔案')
if unmatched_files:
    print(f'  無記錄檔案:')
    for f in unmatched_files:
        print(f'    - {f}')
